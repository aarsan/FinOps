#!/usr/bin/env python3
"""Report Azure VMs from a billing CSV with effective rate, list rate, and benefits.

The customer's billing CSV (legacy 'Detail_BillingProfile_*.csv' or FOCUS
'FocusCost' export) already contains every VM that was billed in the period:
its resource id, region, the VM size (in the meter name), the rates, and any
benefit attribution. Resource Graph isn't needed.

For each VM, the script reports:
  * VM name, resource group, subscription, region (parsed from resource id)
  * VM size (dominant compute meter)
  * Effective hourly rate (cost / hours), list (PAYG) hourly rate
  * Benefit applied (Reservation / Savings Plan / Spot / Azure Hybrid Benefit
    / negotiated MCA rate / list)
  * Hours billed and total cost in the period
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Hours per "month" used to extrapolate hourly rates → monthly cost.
HOURS_PER_MONTH = 730


# ---------- CSV schemas ----------

LEGACY_VM_SCHEMA = {
    "name":             "legacy",
    "category":         "meterCategory",
    "category_value":   "Virtual Machines",
    "subcategory":      "meterSubCategory",
    "meter":            "meterName",
    "region":           "resourceLocation",
    "unit":             "unitOfMeasure",
    "unit_pattern":     re.compile(r"^\s*1\s*Hour\s*$"),
    "effective_price":  "effectivePrice",
    "list_price":       "payGPrice",
    "contracted_price": "unitPrice",
    "currency":         "pricingCurrency",
    "pricing_model":    "pricingModel",
    "benefit_name":     "benefitName",
    "benefit_id":       "benefitId",
    "benefit_type":     None,                # legacy infers from pricingModel
    "resource_id":      "resourceId",
    "quantity":         "quantity",
    "cost":             "costInBillingCurrency",
    "list_cost":        "paygCostInBillingCurrency",
    "period_start":     "billingPeriodStartDate",
    "period_end":       "billingPeriodEndDate",
}
FOCUS_VM_SCHEMA = {
    "name":             "FOCUS",
    "category":         "x_SkuMeterCategory",   # 'Virtual Machines'
    "category_value":   "Virtual Machines",
    "subcategory":      "x_SkuMeterSubcategory",
    "meter":            "x_SkuMeterName",
    "region":           "RegionId",
    "unit":             "PricingUnit",
    "unit_pattern":     re.compile(r"^\s*Hours?\s*$|^\s*1\s*Hour\s*$"),
    "effective_price":  "x_EffectiveUnitPrice",
    "list_price":       "ListUnitPrice",
    "contracted_price": "ContractedUnitPrice",
    "currency":         "BillingCurrency",
    "pricing_model":    "PricingCategory",      # Standard | Committed | DynamicPricing
    "benefit_name":     "CommitmentDiscountName",
    "benefit_id":       "CommitmentDiscountId",
    "benefit_type":     "CommitmentDiscountType",
    "resource_id":      "ResourceId",
    "quantity":         "PricingQuantity",
    "cost":             "EffectiveCost",
    "list_cost":        "ListCost",
    "period_start":     "BillingPeriodStart",
    "period_end":       "BillingPeriodEnd",
}


def _detect_schema(fieldnames: list[str]) -> Optional[dict]:
    """Pick the schema whose required columns are all present."""
    for schema in (LEGACY_VM_SCHEMA, FOCUS_VM_SCHEMA):
        cols = {schema[k] for k in (
            "category", "subcategory", "meter", "region", "unit",
            "effective_price", "list_price", "currency",
            "resource_id", "quantity", "cost",
        )}
        if cols.issubset(fieldnames):
            return schema
    return None


def _to_float(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


# ---------- VM aggregation from CSV ----------

@dataclass
class VmBilling:
    resource_id: str
    name: str
    resource_group: str
    subscription_id: str
    location: str = ""
    quantity_total:  float = 0.0
    cost_total:      float = 0.0
    list_cost_total: float = 0.0
    # Per-row savings: only sum (list - effective) for rows where ListCost > 0.
    # Avoids the FOCUS quirk where reservation amortization or fee rows have
    # EffectiveCost populated but no ListCost equivalent.
    savings_total:   float = 0.0
    savings_list:    float = 0.0   # list portion of savings_total (for sanity)
    currency: str = ""
    # Cost (in billing currency) per (meter, pricingModel, benefitName) — used
    # to pick the dominant meter/benefit when a VM has mixed billing.
    meter_buckets:   dict[tuple, float] = field(
        default_factory=lambda: defaultdict(float))
    # Cost per benefit name (so the dominant benefit reflects spend, not hours).
    benefit_cost:    dict[str, float] = field(
        default_factory=lambda: defaultdict(float))
    # Type per benefit name (Reservation / SavingsPlan), so we can summarize.
    benefit_types:   dict[str, str] = field(default_factory=dict)
    # First-seen sample of the COMPUTE meter rates (we ignore Windows surcharge
    # rows when picking samples, so the rates here are for the base VM).
    sample_effective:  Optional[float] = None
    sample_list:       Optional[float] = None
    sample_contracted: Optional[float] = None
    sample_pricing_model: Optional[str] = None
    # Windows-license-surcharge bookkeeping (separate meter rows)
    has_windows_surcharge: bool = False
    windows_cost: float = 0.0
    windows_hours: float = 0.0


def _parse_vm_resource_id(rid: str) -> Optional[tuple[str, str, str, str]]:
    """Returns (subscriptionId, resourceGroup, name, lowered_id) or None."""
    if not rid:
        return None
    parts = rid.split("/")
    # /subscriptions/{sub}/resourceGroups/{rg}/providers/Microsoft.Compute/virtualMachines/{name}
    if len(parts) < 9:
        return None
    if parts[1].lower() != "subscriptions" or parts[3].lower() != "resourcegroups":
        return None
    if parts[5].lower() != "providers" or parts[6].lower() != "microsoft.compute":
        return None
    if parts[7].lower() != "virtualmachines":
        return None
    return (parts[2], parts[4], parts[8], rid.lower())


def aggregate_vm_billing(path: Path) -> tuple[dict[str, VmBilling], dict]:
    """Stream the CSV; aggregate billing rows by VM resourceId.

    Pre-filter keeps only lines containing ',Virtual Machines,' which matches
    the meterCategory / x_SkuMeterCategory column value. The DictReader stage
    then enforces the precise filters (right column, hourly meter, has a
    /virtualMachines/ resource id).

    Returns (aggregations_by_resource_id, period_info).
    period_info = {"start": str|None, "end": str|None, "days": int|None,
                   "label": str, "schema": str}
    """
    size_mb = path.stat().st_size / (1024 * 1024)
    print(f"Loading VM billing from '{path}' ({size_mb:,.1f} MB)...")
    t0 = time.perf_counter()

    candidate_lines: list[str] = []
    rows_scanned   = 0
    progress_every = 250_000
    needle = ",Virtual Machines,"

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        header_line = f.readline()
        if not header_line:
            raise RuntimeError("Usage CSV is empty.")
        candidate_lines.append(header_line)
        for line in f:
            rows_scanned += 1
            if needle in line:
                candidate_lines.append(line)
            if rows_scanned % progress_every == 0:
                elapsed = time.perf_counter() - t0
                print(f"  scanned {rows_scanned:,} rows, "
                      f"{len(candidate_lines) - 1:,} VM-candidates so far "
                      f"({elapsed:.1f}s)...")

    elapsed = time.perf_counter() - t0
    print(f"  pre-filter complete: {rows_scanned:,} rows scanned, "
          f"{len(candidate_lines) - 1:,} candidates kept in {elapsed:.1f}s.")

    reader = csv.DictReader(candidate_lines)
    schema = _detect_schema(list(reader.fieldnames or []))
    if schema is None:
        raise RuntimeError(
            "Usage CSV is neither a legacy 'Detail_BillingProfile_*' nor a "
            "FOCUS export. Required columns not found."
        )
    print(f"  detected schema: {schema['name']}")

    cat_col       = schema["category"]
    cat_value     = schema["category_value"]
    sub_col       = schema["subcategory"]
    unit_col      = schema["unit"]
    unit_re       = schema["unit_pattern"]
    rid_col       = schema["resource_id"]
    meter_col     = schema["meter"]
    reg_col       = schema["region"]
    qty_col       = schema["quantity"]
    cost_col      = schema["cost"]
    list_cost_col = schema["list_cost"]
    eff_col       = schema["effective_price"]
    list_col      = schema["list_price"]
    contract_col  = schema.get("contracted_price")
    cur_col       = schema["currency"]
    pm_col        = schema.get("pricing_model")
    bn_col        = schema.get("benefit_name")
    bt_col        = schema.get("benefit_type")
    ps_col        = schema.get("period_start")
    pe_col        = schema.get("period_end")

    # Some optional columns may be absent on older exports — defensively check.
    fields = set(reader.fieldnames or [])
    if contract_col not in fields: contract_col = None
    if pm_col       not in fields: pm_col       = None
    if bn_col       not in fields: bn_col       = None
    if bt_col       not in fields: bt_col       = None
    if ps_col       not in fields: ps_col       = None
    if pe_col       not in fields: pe_col       = None

    aggregations: dict[str, VmBilling] = {}
    matched = 0
    skipped_non_vm = 0
    period_start_str: Optional[str] = None
    period_end_str:   Optional[str] = None

    for row in reader:
        if row.get(cat_col) != cat_value:
            continue
        if not unit_re.match(row.get(unit_col) or ""):
            continue

        # Capture period dates from the first matching row; FOCUS uses ISO
        # timestamps, legacy uses MM/DD/YYYY.
        if period_start_str is None and ps_col:
            v = (row.get(ps_col) or "").strip()
            if v:
                period_start_str = v
        if period_end_str is None and pe_col:
            v = (row.get(pe_col) or "").strip()
            if v:
                period_end_str = v

        rid = row.get(rid_col) or ""
        parsed = _parse_vm_resource_id(rid)
        if parsed is None:
            # Could be VMSS, ARO, AKS-managed VMs, etc.
            skipped_non_vm += 1
            continue
        sub_id, rg, name, rid_lower = parsed

        matched += 1
        agg = aggregations.get(rid_lower)
        if agg is None:
            agg = VmBilling(
                resource_id=rid,
                name=name,
                resource_group=rg,
                subscription_id=sub_id,
            )
            aggregations[rid_lower] = agg

        if not agg.location:
            agg.location = row.get(reg_col) or ""

        qty       = _to_float(row.get(qty_col))       or 0.0
        cost      = _to_float(row.get(cost_col))      or 0.0
        list_cost = _to_float(row.get(list_cost_col)) or 0.0

        meter = row.get(meter_col) or ""
        sub_cat = row.get(sub_col) or ""
        # Detect the Windows-license surcharge: separate meter row whose
        # subcategory ends in " Windows" (e.g. 'Virtual Machines Dv3 Series
        # Windows'). Track it but DO NOT use it as the dominant compute meter.
        is_windows_surcharge = sub_cat.endswith(" Windows")

        agg.quantity_total  += qty
        agg.cost_total      += cost
        agg.list_cost_total += list_cost
        # Only credit savings on rows where both costs exist. This is the right
        # math: per-row (list − effective). Aggregate ratios are misleading
        # because some rows (e.g. RI true-up / fees) have EffectiveCost only.
        if list_cost > 0 and cost <= list_cost + 1e-9:
            agg.savings_total += (list_cost - cost)
            agg.savings_list  += list_cost

        if is_windows_surcharge:
            agg.has_windows_surcharge = True
            agg.windows_cost  += cost
            agg.windows_hours += qty
            continue  # don't pollute the compute-meter pool

        pm = (row.get(pm_col)   or "") if pm_col else ""
        bn = (row.get(bn_col)   or "") if bn_col else ""
        bt = (row.get(bt_col)   or "") if bt_col else ""
        agg.meter_buckets[(meter, pm, bn)] += cost
        if bn:
            agg.benefit_cost[bn] += cost
            if bt and bn not in agg.benefit_types:
                agg.benefit_types[bn] = bt

        if agg.sample_effective is None:
            agg.sample_effective = _to_float(row.get(eff_col))
        if agg.sample_list is None:
            agg.sample_list = _to_float(row.get(list_col))
        if agg.sample_contracted is None and contract_col:
            agg.sample_contracted = _to_float(row.get(contract_col))
        if not agg.currency:
            agg.currency = row.get(cur_col) or ""
        if agg.sample_pricing_model is None and pm:
            agg.sample_pricing_model = pm

    elapsed = time.perf_counter() - t0
    print(f"Aggregated {len(aggregations)} VMs from {matched} matching billing "
          f"rows ({rows_scanned:,} total rows scanned in {elapsed:.1f}s).")
    if skipped_non_vm:
        print(f"  (skipped {skipped_non_vm} non-VM rows that hit the pre-filter "
              "— scale-set instances, etc.)")

    period_info = _build_period_info(period_start_str, period_end_str,
                                     schema_name=schema["name"])
    return (aggregations, period_info)


def _build_period_info(start: Optional[str], end: Optional[str],
                       *, schema_name: str) -> dict:
    """Parse start/end strings into a period descriptor.

    Accepts ISO timestamps (FOCUS), MM/DD/YYYY (legacy), or 'YYYY-MM-DD'.
    Returns dict with start, end (date-only strings), days, label.
    """
    from datetime import datetime, timezone

    def _parse(s: Optional[str]):
        if not s:
            return None
        s2 = s.split("T")[0].strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s2, fmt).date()
            except ValueError:
                pass
        return None

    sd = _parse(start)
    ed = _parse(end)
    days: Optional[int] = None
    if sd and ed:
        # FOCUS BillingPeriodEnd is exclusive (next month start); legacy is
        # inclusive end-of-month. Treat ed as exclusive if equal to first of
        # next month, else inclusive +1.
        diff = (ed - sd).days
        if diff <= 0:
            days = None
        else:
            # Heuristic: if ed.day == 1 and ed > sd, treat as exclusive end.
            days = diff if ed.day == 1 else diff + 1

    if sd and ed and days:
        label = f"{sd.isoformat()} → {ed.isoformat()} ({days} days)"
    elif sd and ed:
        label = f"{sd.isoformat()} → {ed.isoformat()}"
    elif sd or ed:
        label = (sd or ed).isoformat() if (sd or ed) else "unknown"
    else:
        label = "unknown period"

    return {
        "start":  sd.isoformat() if sd else None,
        "end":    ed.isoformat() if ed else None,
        "days":   days,
        "label":  label,
        "schema": schema_name,
    }


# ---------- Build per-VM result rows ----------

def _dominant_meter(agg: VmBilling) -> tuple[str, str, str]:
    """Pick the meter/pricingModel/benefit (by cost) for the VM."""
    if not agg.meter_buckets:
        return ("", "", "")
    return max(agg.meter_buckets, key=agg.meter_buckets.get)


def _classify_benefit(agg: VmBilling) -> tuple[str, str, str]:
    """Returns (benefit_label, benefit_category, ahb_status).

    benefit_label    — detailed: 'Reservation: <name> (+N more)', 'Savings Plan: <name>', etc.
    benefit_category — coarse:   'Reservation' / 'Reservation (multiple)' / 'Savings Plan' /
                                 'Spot' / 'MCA negotiated' / 'Negotiated' / 'List'
    ahb_status       — 'No (Windows surcharge billed)' / 'N/A or AHB applied'
    """

    # AHB: if a Windows surcharge appears, AHB is definitely OFF for that VM.
    # Otherwise we can't distinguish "Linux" from "Windows w/ AHB" from billing
    # alone — but the absence of a surcharge tells us they're paying compute
    # rates only, which is the desired state either way.
    if agg.has_windows_surcharge:
        ahb = "No (Windows surcharge billed)"
    else:
        ahb = "N/A or AHB applied"

    if agg.benefit_cost:
        ranked = sorted(agg.benefit_cost.items(), key=lambda kv: -kv[1])
        primary_name, _ = ranked[0]
        primary_type = agg.benefit_types.get(primary_name) \
            or agg.sample_pricing_model or "Benefit"
        type_label = {
            "Reservation": "Reservation",
            "SavingsPlan": "Savings Plan",
            "Committed":   "Reservation/Savings Plan",
        }.get(primary_type, primary_type)

        label = f"{type_label}: {primary_name}"
        if len(ranked) > 1:
            label = f"{label} (+{len(ranked) - 1} more)"

        # Coarse category — what the customer actually wants to see in a summary
        if primary_type == "Reservation":
            category = "Reservation (multiple)" if len(ranked) > 1 else "Reservation"
        elif primary_type == "SavingsPlan":
            category = "Savings Plan (multiple)" if len(ranked) > 1 else "Savings Plan"
        else:
            category = type_label
        return (label, category, ahb)

    pm = (agg.sample_pricing_model or "").lower()
    if pm == "spot":
        return ("Spot", "Spot", ahb)

    if (agg.sample_contracted is not None and agg.sample_list is not None
            and agg.sample_contracted < agg.sample_list - 1e-9):
        return ("MCA negotiated rate", "MCA negotiated", ahb)
    if (agg.sample_effective is not None and agg.sample_list is not None
            and agg.sample_effective < agg.sample_list - 1e-9):
        return ("Negotiated rate", "Negotiated", ahb)
    return ("List price", "List", ahb)


def build_results(billing: dict[str, VmBilling]) -> list[dict]:
    out: list[dict] = []
    for agg in billing.values():
        meter_tuple = _dominant_meter(agg)
        meter = meter_tuple[0] or None

        # Compute hours/costs are everything minus the Windows surcharge bucket.
        compute_hours    = max(agg.quantity_total  - agg.windows_hours, 0.0)
        compute_cost     = max(agg.cost_total      - agg.windows_cost,  0.0)
        compute_list     = agg.savings_list  # list_cost only on comparable rows
        compute_savings  = agg.savings_total # per-row sum of (list - effective)

        effective_rate = list_rate = contract_rate = None
        if compute_hours > 0:
            effective_rate = round(compute_cost / compute_hours, 6)
            # Prefer the per-row sample rate over the aggregate ratio: it's
            # the actual list price for this SKU, not a blended approximation.
            if agg.sample_list is not None:
                list_rate = round(agg.sample_list, 6)
        contract_rate = round(agg.sample_contracted, 6) \
            if agg.sample_contracted is not None else None

        monthly_cost = monthly_list = monthly_savings = discount_pct = None
        if effective_rate is not None:
            monthly_cost = round(effective_rate * HOURS_PER_MONTH, 2)
            if list_rate is not None:
                monthly_list    = round(list_rate * HOURS_PER_MONTH, 2)
                monthly_savings = round(monthly_list - monthly_cost, 2)
                if list_rate > 0:
                    discount_pct = round(
                        (list_rate - effective_rate) / list_rate * 100, 2)

        benefit, benefit_category, ahb = _classify_benefit(agg)

        out.append({
            "Name":                 agg.name,
            "ResourceGroup":        agg.resource_group,
            "SubscriptionId":       agg.subscription_id,
            "Location":             agg.location,
            "VmSize":               meter,
            "Currency":             agg.currency or None,
            "EffectiveHourlyRate":  effective_rate,
            "ContractedHourlyRate": contract_rate,
            "PayGHourlyRate":       list_rate,
            "BenefitCategory":      benefit_category,
            "Benefit":              benefit,
            "AhbStatus":            ahb,
            "DiscountPercent":      discount_pct,
            "BillingHours":         round(compute_hours, 4) if compute_hours else None,
            "ActualCostInPeriod":   round(compute_cost, 4)  if compute_cost  else None,
            "WindowsSurchargeCost": round(agg.windows_cost, 4) if agg.windows_cost else None,
            "ActualListInPeriod":   round(compute_list, 4)  if compute_list  else None,
            "ActualSavings":        round(compute_savings, 4) if compute_savings else None,
            "EstMonthlyCost":       monthly_cost,
            "EstMonthlyListCost":   monthly_list,
            "EstMonthlySavings":    monthly_savings,
            "ResourceId":           agg.resource_id,
        })
    return out


# ---------- Insights / recommendations (callable from dashboard) ----------

def compute_vm_by_category(results: list[dict]) -> dict[str, tuple[int, float]]:
    """Group VMs by BenefitCategory: returns {category: (count, total_cost)}."""
    out: dict[str, tuple[int, float]] = defaultdict(lambda: (0, 0.0))
    for r in results:
        cat = r["BenefitCategory"] or "Unknown"
        n, c = out[cat]
        out[cat] = (n + 1, c + (r["ActualCostInPeriod"] or 0))
    return dict(out)


def compute_observed_ri_discount(results: list[dict]) -> Optional[float]:
    """Cost-weighted average RI discount derived from existing RI-covered VMs.
    Returns 0.0–0.95 or None if no RI-covered VMs to learn from."""
    ri_covered = [r for r in results
                  if (r["BenefitCategory"] in ("Reservation", "Reservation (multiple)"))
                  and r["EffectiveHourlyRate"] is not None
                  and r["PayGHourlyRate"] is not None
                  and r["PayGHourlyRate"] > 0]
    if not ri_covered:
        return None
    weighted_sum = sum(
        (1 - (r["EffectiveHourlyRate"] / r["PayGHourlyRate"]))
        * (r["ActualCostInPeriod"] or 0)
        for r in ri_covered
    )
    weight_total = sum(r["ActualCostInPeriod"] or 0 for r in ri_covered)
    if weight_total <= 0:
        return None
    return max(0.0, min(0.95, weighted_sum / weight_total))


def compute_ri_candidates(results: list[dict],
                          observed_ri_discount: Optional[float],
                          annualize: float) -> list[dict]:
    """Group on-demand VMs by (VmSize, Location). Returns the meaningful
    subset (always-on, or ≥ 1 month cumulative hours), sorted by projected
    annual savings descending. Caps at 25 rows."""
    candidate_vms = [r for r in results
                     if r["BenefitCategory"] in ("MCA negotiated", "Negotiated", "List")
                     and r["VmSize"]
                     and (r["BillingHours"] or 0) > 0]
    groups: dict[tuple[str, str], dict] = {}
    for r in candidate_vms:
        key = (r["VmSize"], r["Location"] or "")
        agg = groups.setdefault(key, {
            "VmSize":     r["VmSize"],
            "Location":   r["Location"],
            "VmCount":    0,
            "TotalHours": 0.0,
            "ActualCost": 0.0,
            "ListCost":   0.0,
            "Currency":   r["Currency"] or "",
        })
        agg["VmCount"]    += 1
        agg["TotalHours"] += r["BillingHours"] or 0
        agg["ActualCost"] += r["ActualCostInPeriod"] or 0
        agg["ListCost"]   += r["ActualListInPeriod"] or 0

    ALWAYS_ON_HOURS = 0.80 * HOURS_PER_MONTH
    rows: list[dict] = []
    for agg in groups.values():
        hrs_per_vm = agg["TotalHours"] / max(agg["VmCount"], 1)
        coverage = ("Always-on" if hrs_per_vm >= ALWAYS_ON_HOURS
                    else "Mostly on" if hrs_per_vm >= 0.5 * HOURS_PER_MONTH
                    else "Bursty / dev")
        proj_period = (agg["ActualCost"] * observed_ri_discount) \
            if observed_ri_discount is not None else None
        proj_annual = proj_period * annualize if proj_period is not None else None
        rows.append({
            **agg,
            "AvgHoursPerVm":          round(hrs_per_vm, 1),
            "Coverage":               coverage,
            "ProjectedSavingsPeriod": round(proj_period, 2) if proj_period is not None else None,
            "ProjectedSavingsAnnual": round(proj_annual, 2) if proj_annual is not None else None,
        })
    rows.sort(key=lambda x: -(x.get("ProjectedSavingsAnnual") or x["ActualCost"]))
    return [x for x in rows
            if x["Coverage"] == "Always-on" or x["TotalHours"] >= HOURS_PER_MONTH][:25]


# Public reference rate for Azure Spot — typical 60-90% off list, region/SKU
# dependent. We use 0.60 as a conservative defensible estimate so the
# 'opportunity' number is realistic, not aspirational.
SPOT_REFERENCE_DISCOUNT = 0.60


def compute_spot_candidates(results: list[dict]) -> list[dict]:
    """Group on-demand 'Bursty / dev' VMs (low avg hours/VM) by (VmSize, Location).

    These are the natural Spot candidates: workloads that already run < 50% of
    the period are typically dev/test or batch — fault-tolerant enough that
    Spot evictions are tolerable. Always-on VMs are NOT Spot candidates (RI
    or SP fits better).

    Returns rows sorted by projected period savings descending. Caps at 25.
    """
    candidate_vms = [r for r in results
                     if r["BenefitCategory"] in ("MCA negotiated", "Negotiated", "List")
                     and r["VmSize"]
                     and (r["BillingHours"] or 0) > 0]
    groups: dict[tuple[str, str], dict] = {}
    for r in candidate_vms:
        key = (r["VmSize"], r["Location"] or "")
        agg = groups.setdefault(key, {
            "VmSize":     r["VmSize"],
            "Location":   r["Location"],
            "VmCount":    0,
            "TotalHours": 0.0,
            "ActualCost": 0.0,
            "Currency":   r["Currency"] or "",
        })
        agg["VmCount"]    += 1
        agg["TotalHours"] += r["BillingHours"] or 0
        agg["ActualCost"] += r["ActualCostInPeriod"] or 0

    BURSTY_HOURS_THRESHOLD = 0.50 * HOURS_PER_MONTH
    rows: list[dict] = []
    for agg in groups.values():
        hrs_per_vm = agg["TotalHours"] / max(agg["VmCount"], 1)
        if hrs_per_vm >= BURSTY_HOURS_THRESHOLD:
            continue   # not bursty enough to qualify as Spot-friendly
        proj_period = agg["ActualCost"] * SPOT_REFERENCE_DISCOUNT
        rows.append({
            **agg,
            "AvgHoursPerVm":          round(hrs_per_vm, 1),
            "Coverage":               "Bursty / dev",
            "ProjectedSavingsPeriod": round(proj_period, 2),
        })
    rows.sort(key=lambda x: -(x["ProjectedSavingsPeriod"] or 0))
    # Only return groups with at least 1 month of cumulative hours so we
    # don't surface tiny one-off VMs.
    return [x for x in rows if x["TotalHours"] >= HOURS_PER_MONTH][:25]


def compute_ahb_vm_candidates(results: list[dict]) -> list[dict]:
    """Per-VM list where Windows-license surcharge was billed (AHB not applied)."""
    cands = [r for r in results
             if r["AhbStatus"] == "No (Windows surcharge billed)"]
    cands.sort(key=lambda r: -(r["WindowsSurchargeCost"] or 0))
    return cands


# ---------- CSV resolution ----------

def find_default_usage_csv(script_path: Path) -> Optional[Path]:
    workspace = script_path.parent.parent
    data_dir  = workspace / "data"
    if not data_dir.is_dir():
        return None
    csvs = sorted(
        data_dir.glob("*.csv"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not csvs:
        return None
    if len(csvs) == 1:
        return csvs[0]
    legacy = [p for p in csvs if p.name.startswith("Detail_BillingProfile_")]
    return legacy[0] if legacy else csvs[0]


# ---------- main ----------

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--usage-csv",
        help="Detailed billing CSV. If omitted, the script picks the single "
             "CSV under <workspace>/data/ (or the most recent if there are "
             "multiple).",
    )
    parser.add_argument(
        "--export-csv",
        default="reports/vm-report.csv",
        help="Output CSV path (default: %(default)s).",
    )
    parser.add_argument(
        "--export-xlsx",
        default="reports/vm-report.xlsx",
        help="Output Excel path (default: %(default)s). "
             "Pass an empty string to skip.",
    )
    args = parser.parse_args(argv)

    script_path = Path(__file__).resolve()

    # ----- Resolve usage CSV
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = find_default_usage_csv(script_path)
        if not found:
            print("\nERROR: No usage CSV provided and no *.csv found in "
                  "<workspace>/data/. Pass --usage-csv.", file=sys.stderr)
            return 2
        usage_csv = found.resolve()
        print(f"Auto-detected usage CSV: {usage_csv}")

    if not usage_csv.exists():
        print(f"\nERROR: Usage CSV not found: {usage_csv}", file=sys.stderr)
        return 2

    # ----- Aggregate billing per VM
    billing, period = aggregate_vm_billing(usage_csv)
    if not billing:
        print("\nNo VM billing rows found in the CSV.")
        return 0

    results = build_results(billing)
    results.sort(key=lambda r: -(r["ActualCostInPeriod"] or 0))

    total_period_cost      = sum(r["ActualCostInPeriod"]   or 0 for r in results)
    total_period_list      = sum(r["ActualListInPeriod"]   or 0 for r in results)
    total_period_savings   = sum(r["ActualSavings"]        or 0 for r in results)
    total_windows_surcharge= sum(r["WindowsSurchargeCost"] or 0 for r in results)
    total_monthly_cost     = sum(r["EstMonthlyCost"]       or 0 for r in results)
    total_monthly_list     = sum(r["EstMonthlyListCost"]   or 0 for r in results)
    total_monthly_save     = sum(r["EstMonthlySavings"]    or 0 for r in results)
    cur = next((r["Currency"] for r in results if r["Currency"]), "")

    # Annualization factor: scale period totals → 12-month projection.
    period_days = period.get("days") or 30
    annualize   = 365.0 / period_days

    print()
    print(f"Billing period                              : {period['label']}")
    print(f"Total VMs (billed in period)                : {len(results)}")
    print(f"Actual compute cost in period               : {total_period_cost:,.2f} {cur}")
    print(f"Actual list (PAYG) cost in period           : {total_period_list:,.2f} {cur}")
    print(f"Actual savings in period (per-row)          : {total_period_savings:,.2f} {cur}")
    print(f"Windows-license surcharge billed in period  : {total_windows_surcharge:,.2f} {cur}")
    print(f"Annualized actual cost (period × 365/days)  : {total_period_cost * annualize:,.2f} {cur}/year")
    print(f"Annualized actual savings vs list           : {total_period_savings * annualize:,.2f} {cur}/year")
    print(f"Run-rate monthly cost (rate × {HOURS_PER_MONTH}h)         : {total_monthly_cost:,.2f} {cur}/month")
    print(f"Run-rate monthly list cost                  : {total_monthly_list:,.2f} {cur}/month")
    print(f"Run-rate monthly savings vs list            : {total_monthly_save:,.2f} {cur}/month")

    # Coarse breakdown by benefit category (for the headline summary).
    by_category = compute_vm_by_category(results)
    if by_category:
        print("\nActual compute cost by benefit category:")
        for cat, (n, c) in sorted(by_category.items(), key=lambda kv: -kv[1][1]):
            print(f"  {cat:<28} {n:>5} VMs  {c:>14,.2f} {cur}")

    # Detailed benefit breakdown (full RI/SP names) — written to the workbook
    # only; too noisy for the console.
    by_benefit: dict[str, float] = defaultdict(float)
    for r in results:
        by_benefit[r["Benefit"]] += r["ActualCostInPeriod"] or 0

    # AHB callout: VMs where a Windows surcharge was billed (i.e., AHB not used).
    ahb_candidates = compute_ahb_vm_candidates(results)
    total_windows_candidates = sum(r["WindowsSurchargeCost"] or 0
                                   for r in ahb_candidates)
    if ahb_candidates:
        ann_surcharge = total_windows_candidates * annualize
        print(f"\nAzure Hybrid Benefit opportunity: {len(ahb_candidates)} VM(s) "
              f"paid {total_windows_candidates:,.2f} {cur} in Windows-license "
              f"surcharge in this period (~{ann_surcharge:,.2f} {cur}/year).")
        print(f"  Top {min(10, len(ahb_candidates))} by surcharge cost (period):")
        for r in ahb_candidates[:10]:
            print(f"    {r['Name']:<35} {r['VmSize'] or '':<25} "
                  f"{(r['WindowsSurchargeCost'] or 0):>10,.2f} {cur}")
        if len(ahb_candidates) > 10:
            remaining = sum(r["WindowsSurchargeCost"] or 0
                            for r in ahb_candidates[10:])
            print(f"    ... and {len(ahb_candidates) - 10} more "
                  f"({remaining:,.2f} {cur})")

    # ----- RI candidate analysis (uses the customer's observed RI discount)
    observed_ri_discount = compute_observed_ri_discount(results)
    ri_top = compute_ri_candidates(results, observed_ri_discount, annualize)

    if observed_ri_discount is not None and ri_top:
        total_ann = sum(x.get("ProjectedSavingsAnnual") or 0 for x in ri_top)
        print(f"\nReservation candidates "
              f"(observed customer RI discount: {observed_ri_discount * 100:.1f}% off list)")
        print(f"  Top {len(ri_top)} (VmSize × region) groups currently on-demand "
              f"(projected ~{total_ann:,.0f} {cur}/year if covered):")
        print(f"  {'VmSize':<22} {'Region':<14} {'VMs':>4} {'Hrs/VM':>8}  "
              f"{'PeriodCost':>12}  {'~PeriodSav':>11}  {'~AnnualSav':>11}  Coverage")
        print("  " + "-" * 110)
        for x in ri_top:
            ps = x["ProjectedSavingsPeriod"] or 0
            pa = x["ProjectedSavingsAnnual"] or 0
            print(f"  {x['VmSize'][:22]:<22} {x['Location']:<14} "
                  f"{x['VmCount']:>4} {x['AvgHoursPerVm']:>8.1f}  "
                  f"{x['ActualCost']:>12,.2f}  {ps:>11,.2f}  {pa:>11,.2f}  "
                  f"{x['Coverage']}")
    elif candidate_vms and observed_ri_discount is None:
        print(f"\nReservation candidates: {len(candidate_vms)} on-demand VM(s) "
              "found, but no existing RI-covered VMs to derive a discount rate "
              "from. Cannot project savings.")

    # ----- Export
    out_path = Path(args.export_csv).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)
    print(f"\nReport exported to {out_path}")

    if args.export_xlsx:
        xlsx_path = Path(args.export_xlsx).resolve()
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            write_xlsx(results, xlsx_path, currency=cur,
                       period=period, annualize=annualize,
                       total_period_cost=total_period_cost,
                       total_period_list=total_period_list,
                       total_period_savings=total_period_savings,
                       total_windows_surcharge=total_windows_surcharge,
                       total_monthly_cost=total_monthly_cost,
                       total_monthly_list=total_monthly_list,
                       total_monthly_save=total_monthly_save,
                       by_category=by_category,
                       by_benefit=by_benefit,
                       ahb_candidates=ahb_candidates,
                       ri_candidates=ri_top,
                       observed_ri_discount=observed_ri_discount)
            print(f"Excel report exported to {xlsx_path}")
        except ImportError:
            print("WARNING: openpyxl not installed; skipping Excel export.",
                  file=sys.stderr)
        except PermissionError:
            print(f"WARNING: cannot write {xlsx_path} — is it open in Excel? "
                  f"Close it or pass --export-xlsx <other-path>.",
                  file=sys.stderr)

    return 0


def write_xlsx(results, out_path, *, currency, period, annualize,
               total_period_cost,
               total_period_list, total_period_savings,
               total_windows_surcharge,
               total_monthly_cost, total_monthly_list, total_monthly_save,
               by_category, by_benefit, ahb_candidates,
               ri_candidates, observed_ri_discount) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "VMs"

    bold        = Font(bold=True)
    bold_white  = Font(bold=True, color="FFFFFF")
    big_bold    = Font(bold=True, size=14)
    accent      = Font(bold=True, color="0070C0")
    header_fill = PatternFill("solid", fgColor="305496")
    money_fmt   = "#,##0.00"
    rate_fmt    = "#,##0.000000"
    pct_fmt     = '0.00"%"'

    period_label = period.get("label") or "unknown period"

    ws["A1"] = (f"Virtual Machines — {len(results)} VMs billed | "
                f"Period: {period_label}")
    ws["A1"].font = big_bold
    ws.merge_cells("A1:D1")

    ws["A3"] = "Actual compute cost (period)";          ws["B3"] = round(total_period_cost, 2);            ws["C3"] = currency
    ws["A4"] = "Actual list (PAYG) cost (period)";      ws["B4"] = round(total_period_list, 2);            ws["C4"] = currency
    ws["A5"] = "Actual savings vs list (period)";       ws["B5"] = round(total_period_savings, 2);         ws["C5"] = currency
    ws["A6"] = "Windows-license surcharge (period)";    ws["B6"] = round(total_windows_surcharge, 2);      ws["C6"] = currency
    ws["A7"] = "Annualized actual cost (period × 365/days)";   ws["B7"] = round(total_period_cost * annualize, 2); ws["C7"] = f"{currency}/year"
    ws["A8"] = "Annualized actual savings vs list";     ws["B8"] = round(total_period_savings * annualize, 2); ws["C8"] = f"{currency}/year"
    ws["A9"] = f"Run-rate monthly cost (rate × {HOURS_PER_MONTH}h)";  ws["B9"]  = round(total_monthly_cost, 2);  ws["C9"]  = f"{currency}/month"
    ws["A10"] = "Run-rate monthly list cost";           ws["B10"] = round(total_monthly_list, 2);           ws["C10"] = f"{currency}/month"
    ws["A11"] = "Run-rate monthly savings vs list";     ws["B11"] = round(total_monthly_save, 2);           ws["C11"] = f"{currency}/month"
    for r in range(3, 12):
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=2).number_format = money_fmt
    # Highlight the most useful headline numbers
    for r in (5, 7, 8, 9):
        ws.cell(row=r, column=1).font = accent
        ws.cell(row=r, column=2).font = accent

    # Coarse breakdown by benefit category
    by_cat_start = 13
    ws.cell(row=by_cat_start, column=1, value="Benefit category").font = bold_white
    ws.cell(row=by_cat_start, column=2, value="VMs").font = bold_white
    ws.cell(row=by_cat_start, column=3, value=f"Actual cost ({currency})").font = bold_white
    for c in (1, 2, 3):
        ws.cell(row=by_cat_start, column=c).fill = header_fill
    for i, (cat, (n, c)) in enumerate(
            sorted(by_category.items(), key=lambda kv: -kv[1][1]),
            start=by_cat_start + 1):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=n)
        ws.cell(row=i, column=3, value=round(c, 2)).number_format = money_fmt
    cat_end = by_cat_start + len(by_category)

    # Detailed benefit breakdown (full RI/SP names) — separate sheet so the
    # main summary stays readable.
    table_start = cat_end + 3

    # Per-VM table on the main sheet
    headers = list(results[0].keys())
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_cols = {"ActualCostInPeriod", "ActualListInPeriod",
                  "ActualSavings", "WindowsSurchargeCost",
                  "EstMonthlyCost", "EstMonthlyListCost", "EstMonthlySavings"}
    rate_cols  = {"EffectiveHourlyRate", "ContractedHourlyRate", "PayGHourlyRate"}
    pct_cols   = {"DiscountPercent"}
    for r_idx, row in enumerate(results, start=table_start + 1):
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            if h in money_cols and row[h] is not None:
                cell.number_format = money_fmt
            elif h in rate_cols and row[h] is not None:
                cell.number_format = rate_fmt
            elif h in pct_cols and row[h] is not None:
                cell.number_format = pct_fmt

    # Column widths
    for c_idx, h in enumerate(headers, start=1):
        letter = get_column_letter(c_idx)
        max_len = len(h)
        for row in results:
            v = row[h]
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max_len + 2
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 50)

    ws.freeze_panes = ws.cell(row=table_start + 1, column=1)
    ws.auto_filter.ref = (
        f"A{table_start}:{get_column_letter(len(headers))}"
        f"{table_start + len(results)}"
    )

    # ----- Detailed benefit breakdown sheet (full RI/SP names)
    if by_benefit:
        bws = wb.create_sheet("Benefit detail")
        bws["A1"] = "Cost by benefit (full reservation/savings-plan names)"
        bws["A1"].font = big_bold
        bws.merge_cells("A1:B1")
        bws.cell(row=3, column=1, value="Benefit").font = bold_white
        bws.cell(row=3, column=2, value=f"Actual cost ({currency})").font = bold_white
        bws.cell(row=3, column=1).fill = header_fill
        bws.cell(row=3, column=2).fill = header_fill
        for i, (b, c) in enumerate(sorted(by_benefit.items(), key=lambda kv: -kv[1]),
                                   start=4):
            bws.cell(row=i, column=1, value=b)
            bws.cell(row=i, column=2, value=round(c, 2)).number_format = money_fmt
        bws.column_dimensions["A"].width = 80
        bws.column_dimensions["B"].width = 22
        bws.freeze_panes = "A4"

    # ----- AHB opportunity sheet (VMs paying Windows-license surcharge)
    if ahb_candidates:
        aws = wb.create_sheet("AHB opportunity")
        total_surcharge = sum(r["WindowsSurchargeCost"] or 0
                              for r in ahb_candidates)
        aws["A1"] = (f"Azure Hybrid Benefit opportunity — "
                     f"{len(ahb_candidates)} VM(s), "
                     f"{total_surcharge:,.2f} {currency} in surcharge")
        aws["A1"].font = big_bold
        aws.merge_cells("A1:E1")
        ahb_headers = ["Name", "ResourceGroup", "SubscriptionId", "Location",
                       "VmSize", "BillingHours", "WindowsSurchargeCost",
                       "ActualCostInPeriod", "ResourceId"]
        for c_idx, h in enumerate(ahb_headers, start=1):
            cell = aws.cell(row=3, column=c_idx, value=h)
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, r in enumerate(ahb_candidates, start=4):
            for c_idx, h in enumerate(ahb_headers, start=1):
                cell = aws.cell(row=r_idx, column=c_idx, value=r.get(h))
                if h in money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
        # Column widths
        for c_idx, h in enumerate(ahb_headers, start=1):
            letter = get_column_letter(c_idx)
            max_len = len(h)
            for r in ahb_candidates:
                v = r.get(h)
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)), 50))
            aws.column_dimensions[letter].width = max_len + 2
        aws.freeze_panes = "A4"
        aws.auto_filter.ref = (
            f"A3:{get_column_letter(len(ahb_headers))}"
            f"{3 + len(ahb_candidates)}"
        )

    # ----- Recommendations sheet (consolidated insights)
    rws = wb.create_sheet("Recommendations", 0)  # insert as first sheet
    rws["A1"] = f"Optimization recommendations  |  Period: {period_label}"
    rws["A1"].font = big_bold
    rws.merge_cells("A1:E1")

    row = 3

    # ---- AHB ----
    rws.cell(row=row, column=1,
             value="1. Azure Hybrid Benefit (AHB)").font = Font(bold=True, size=12)
    row += 1
    if ahb_candidates:
        ahb_total = sum(r["WindowsSurchargeCost"] or 0 for r in ahb_candidates)
        ahb_annual = ahb_total * annualize
        rws.cell(row=row, column=1,
                 value=f"  • {len(ahb_candidates)} VM(s) paying Windows-license "
                       f"surcharge (no AHB applied).").alignment = Alignment(wrap_text=True)
        row += 1
        rws.cell(row=row, column=1,
                 value=f"  • Surcharge billed this period ({period_label}): "
                       f"{ahb_total:,.2f} {currency}")
        row += 1
        rws.cell(row=row, column=1,
                 value=f"  • Annualized surcharge: ~{ahb_annual:,.2f} {currency}/year")
        row += 1
        rws.cell(row=row, column=1,
                 value="  • Action: enable Azure Hybrid Benefit on these VMs "
                       "(if eligible Windows Server / SQL licenses exist).")
        row += 1
        rws.cell(row=row, column=1,
                 value=f"  • Detail: see 'AHB opportunity' sheet.")
        row += 1
    else:
        rws.cell(row=row, column=1,
                 value="  • No Windows-license surcharge billed in this period. "
                       "AHB is either applied or N/A across all VMs.")
        row += 1
    row += 1

    # ---- Reservations ----
    rws.cell(row=row, column=1,
             value="2. Reservations (RI)").font = Font(bold=True, size=12)
    row += 1
    if observed_ri_discount is not None:
        rws.cell(row=row, column=1,
                 value=f"  • Observed customer RI discount (derived from "
                       f"existing RI-covered VMs in this period): "
                       f"{observed_ri_discount * 100:.1f}% off list.")
        row += 1
    else:
        rws.cell(row=row, column=1,
                 value="  • No existing RI-covered VMs in the period; cannot "
                       "derive a customer-specific RI discount rate.")
        row += 1

    if ri_candidates:
        total_proj_period = sum(c.get("ProjectedSavingsPeriod") or 0 for c in ri_candidates)
        total_proj_annual = sum(c.get("ProjectedSavingsAnnual") or 0 for c in ri_candidates)
        always_on = [c for c in ri_candidates if c["Coverage"] == "Always-on"]
        rws.cell(row=row, column=1,
                 value=f"  • {len(ri_candidates)} (VmSize × region) groups are "
                       f"currently on-demand and would benefit from RIs.")
        row += 1
        rws.cell(row=row, column=1,
                 value=f"  • {len(always_on)} of those run always-on (≥ 80% of "
                       f"hours) and are the strongest candidates.")
        row += 1
        if total_proj_period > 0:
            rws.cell(row=row, column=1,
                     value=f"  • Projected savings this period: "
                           f"~{total_proj_period:,.2f} {currency}.")
            row += 1
            rws.cell(row=row, column=1,
                     value=f"  • Projected ANNUAL savings if these were covered: "
                           f"~{total_proj_annual:,.2f} {currency}/year.")
            rws.cell(row=row, column=1).font = accent
            row += 1
        rws.cell(row=row, column=1,
                 value="  • Action: review the RI-candidate table below; buy "
                       "1-year or 3-year RIs for stable always-on workloads, "
                       "scoped to the right billing scope.")
        row += 1
    else:
        rws.cell(row=row, column=1,
                 value="  • No on-demand VMs that would clearly benefit from "
                       "additional RI coverage based on current run hours.")
        row += 1
    row += 1

    # ---- Savings Plans ----
    rws.cell(row=row, column=1,
             value="3. Compute Savings Plans (SP)").font = Font(bold=True, size=12)
    row += 1
    rws.cell(row=row, column=1, value=(
        "  • Savings Plans are a flatter, more flexible alternative to RIs: "
        "they cover ANY VM family/region (and instance scaling) for a 1- or "
        "3-year hourly commit."
    )).alignment = Alignment(wrap_text=True)
    row += 1
    rws.cell(row=row, column=1, value=(
        "  • Use SPs when the workload mix changes often (size resizing, "
        "family swaps, region migrations); use RIs when the size+region is "
        "stable for the term."
    )).alignment = Alignment(wrap_text=True)
    row += 1
    rws.cell(row=row, column=1, value=(
        "  • Public reference rates (verify against your contract): compute "
        "SP ≈ 17% off list (1-yr) / 33% (3-yr); RIs typically deliver more "
        "for stable workloads."
    )).alignment = Alignment(wrap_text=True)
    row += 1
    rws.cell(row=row, column=1,
             value="  • Action: if the customer has many small-cost on-demand "
                   "VMs across families, an SP simplifies coverage with one "
                   "hourly commit instead of dozens of RIs.")
    row += 2

    # ---- RI candidate table ----
    if ri_candidates:
        rws.cell(row=row, column=1,
                 value="Top RI-candidate (VmSize × region) groups").font = Font(bold=True)
        row += 1
        ri_headers = ["VmSize", "Location", "VmCount", "TotalHours",
                      "AvgHoursPerVm", "Coverage",
                      "ActualCost", "ListCost",
                      "ProjectedSavingsPeriod", "ProjectedSavingsAnnual",
                      "Currency"]
        for c_idx, h in enumerate(ri_headers, start=1):
            cell = rws.cell(row=row, column=c_idx, value=h)
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1
        ri_money_cols = {"ActualCost", "ListCost",
                         "ProjectedSavingsPeriod", "ProjectedSavingsAnnual"}
        for r in ri_candidates:
            for c_idx, h in enumerate(ri_headers, start=1):
                cell = rws.cell(row=row, column=c_idx, value=r.get(h))
                if h in ri_money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
                elif h in {"TotalHours", "AvgHoursPerVm"} and r.get(h) is not None:
                    cell.number_format = "#,##0.0"
            row += 1

    # Width tweaks
    rws.column_dimensions["A"].width = 100
    for c in ("B", "C", "D", "E", "F", "G", "H", "I", "J"):
        rws.column_dimensions[c].width = 18

    wb.save(out_path)


if __name__ == "__main__":
    sys.exit(main())
