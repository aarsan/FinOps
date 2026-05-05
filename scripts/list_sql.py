#!/usr/bin/env python3
"""Scan a billing CSV for SQL Server license cost & Azure Hybrid Benefit (AHB) status.

Covers every place SQL Server licensing shows up in Azure billing:

    1. SQL Server on Azure VMs (IaaS)
       - AHB applied : meter category 'Virtual Machines Licenses',
                       subcategory 'SQL Server Azure Hybrid Benefit'
                       (cost = 0; row exists to attribute the BYOL VM hour).
       - Paid licence: subcategory 'SQL Server Standard' or 'SQL Server
                       Enterprise' — every row here is a saveable license.
       - Free        : 'SQL Server Developer Edition' / 'SQL Server Express
                       Edition' (no charge; flagged for completeness).

    2. Azure SQL Database (PaaS, vCore tier)
       - AHB applied : compute meter only (subcategory ends in
                       '... - Compute Gen5' / Serverless / Hyperscale ...).
                       No matching '... - SQL License' row.
       - Paid licence: a row in subcategory '... - SQL License' (vCore or
                       eDTU). EffectiveCost on this row IS the license fee.

    3. Azure SQL Database (PaaS, DTU tier — Basic / Standard / Premium)
       - DTU bundles compute + license. AHB CANNOT apply on DTU. Customer
         would need to migrate to vCore tier first. Flagged 'Migrate to vCore'.

    4. Azure SQL Managed Instance (PaaS)
       - Same '... SQL License' detection as DB. MI is vCore-only.

For each SQL resource the report shows the license cost in the period, the
compute / storage cost (context), and the potential savings if AHB were
enabled. License cost is the actionable savings number — enabling AHB
removes the SQL License meter entirely (the customer keeps paying compute).
"""
from __future__ import annotations

import argparse
import csv
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Reuse parsing helpers from list_ahb so we don't duplicate schema detection.
sys.path.insert(0, str(Path(__file__).parent))
from list_ahb import (  # noqa: E402
    _detect_schema,
    _to_float,
    _parse_resource_id,
    _build_period_info,
    find_default_usage_csv,
)

HOURS_PER_MONTH = 730


# ---------- Workload classification ----------

@dataclass
class SqlAggregation:
    resource_id: str
    name: str
    resource_group: str
    subscription_id: str
    resource_type: str = ""
    location: str = ""
    currency: str = ""
    workload: str = ""              # 'SQL Server on VM' / 'Azure SQL DB (vCore)' / ...
    tier: str = ""                  # 'Standard' / 'Enterprise' / 'General Purpose' / 'DTU' / ...
    ahb_state: str = ""             # 'Applied' / 'Not applied' / 'N/A (Free)' / 'Not eligible (DTU)'
    # Per-bucket cost: helps explain the determination.
    compute_cost: float = 0.0       # vCore / DTU compute (always charged)
    storage_cost: float = 0.0       # data, backup, LTR storage
    license_cost: float = 0.0       # SQL Server license surcharge (= savings if AHB applied)
    other_cost: float = 0.0         # Defender-for-SQL, etc.
    total_cost: float = 0.0
    quantity_total: float = 0.0
    notes: list[str] = field(default_factory=list)


def aggregate_sql(path: Path) -> tuple[dict[str, SqlAggregation], dict]:
    size_mb = path.stat().st_size / (1024 * 1024)
    print(f"Loading SQL-eligible billing from '{path}' ({size_mb:,.1f} MB)...")
    t0 = time.perf_counter()

    # Pre-filter: any line mentioning SQL or 'Azure SQL'. Cheap substring test.
    KEEP_TOKENS = ("SQL ", ",SQL,", "Azure SQL")
    candidate_lines: list[str] = []
    rows_scanned = 0
    progress_every = 250_000

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        header_line = f.readline()
        if not header_line:
            raise RuntimeError("Usage CSV is empty.")
        candidate_lines.append(header_line)
        for line in f:
            rows_scanned += 1
            if any(t in line for t in KEEP_TOKENS):
                candidate_lines.append(line)
            if rows_scanned % progress_every == 0:
                elapsed = time.perf_counter() - t0
                print(f"  scanned {rows_scanned:,} rows, "
                      f"{len(candidate_lines) - 1:,} candidates so far "
                      f"({elapsed:.1f}s)...")

    elapsed = time.perf_counter() - t0
    print(f"  pre-filter complete: {rows_scanned:,} rows scanned, "
          f"{len(candidate_lines) - 1:,} candidates kept in {elapsed:.1f}s.")

    reader = csv.DictReader(candidate_lines)
    schema = _detect_schema(list(reader.fieldnames or []))
    if schema is None:
        raise RuntimeError(
            "Usage CSV is neither a legacy 'Detail_BillingProfile_*' nor a "
            "FOCUS export. Required columns not found."
        )
    print(f"  detected schema: {schema['name']}")

    cat_col   = schema["category"]
    sub_col   = schema["subcategory"]
    rid_col   = schema["resource_id"]
    qty_col   = schema["quantity"]
    cost_col  = schema["cost"]
    cur_col   = schema["currency"]
    reg_col   = schema["region"]
    ps_col    = schema.get("period_start")
    pe_col    = schema.get("period_end")
    fields = set(reader.fieldnames or [])
    if ps_col not in fields: ps_col = None
    if pe_col not in fields: pe_col = None

    # For Azure SQL DB / MI we infer AHB state from whether any '... SQL License'
    # row exists for that resource in the period. Track per-resource flags.
    sqlres_has_license_meter: dict[str, bool] = {}
    sqlres_has_compute_meter: dict[str, bool] = {}
    # For SQL Server on VM the AHB-applied row has cost=0; the Std/Ent rows
    # have cost>0. We decide "Applied" / "Not applied" / "Partially applied"
    # in a post-pass based on what we observed.
    vmsql_has_ahb_row:    dict[str, bool] = {}
    vmsql_has_paid_lic:   dict[str, bool] = {}
    vmsql_paid_editions:  dict[str, set[str]] = defaultdict(set)

    aggregations: dict[str, SqlAggregation] = {}
    period_start_str: Optional[str] = None
    period_end_str:   Optional[str] = None

    for row in reader:
        cat = row.get(cat_col) or ""
        sub = row.get(sub_col) or ""
        rid = row.get(rid_col) or ""
        if not rid:
            continue

        # Capture period from any matching row.
        if period_start_str is None and ps_col:
            v = (row.get(ps_col) or "").strip()
            if v: period_start_str = v
        if period_end_str is None and pe_col:
            v = (row.get(pe_col) or "").strip()
            if v: period_end_str = v

        # Decide whether this row is SQL-relevant and how to classify it.
        workload = ""
        tier = ""
        ahb_state = ""
        bucket = ""  # which cost bucket: 'compute' / 'license' / 'storage' / 'other'

        # ---- SQL Server on Azure VM (IaaS) ----
        if cat == "Virtual Machines Licenses":
            if sub == "SQL Server Azure Hybrid Benefit":
                workload  = "SQL Server on VM"
                ahb_state = ""          # decided in post-pass
                tier      = ""
                bucket    = "license"   # cost is 0 but log the row
                vmsql_has_ahb_row[rid.lower()] = True
            elif sub == "SQL Server Standard":
                workload  = "SQL Server on VM"
                ahb_state = ""
                tier      = "Standard"
                bucket    = "license"
                vmsql_has_paid_lic[rid.lower()] = True
                vmsql_paid_editions[rid.lower()].add("Standard")
            elif sub == "SQL Server Enterprise":
                workload  = "SQL Server on VM"
                ahb_state = ""
                tier      = "Enterprise"
                bucket    = "license"
                vmsql_has_paid_lic[rid.lower()] = True
                vmsql_paid_editions[rid.lower()].add("Enterprise")
            elif sub == "SQL Server Developer Edition":
                workload  = "SQL Server on VM (free)"
                ahb_state = "N/A (Free edition)"
                tier      = "Developer"
                bucket    = "license"
            elif sub == "SQL Server Express Edition":
                workload  = "SQL Server on VM (free)"
                ahb_state = "N/A (Free edition)"
                tier      = "Express"
                bucket    = "license"
            else:
                # Other VM license rows (e.g. Windows-only) — ignore here.
                continue

        # ---- Azure SQL Database (PaaS) ----
        elif cat == "SQL Database":
            workload = "Azure SQL Database"
            sl = sub.lower()

            # Storage / backup buckets first (don't change AHB state).
            if "storage" in sl or "backup" in sl:
                bucket = "storage"
                tier = _sqldb_tier_from_subcategory(sub)
                # ahb_state inferred at the end from license-meter presence.
            # Compute / license meters
            elif "sql license" in sl:
                bucket = "license"
                tier = _sqldb_tier_from_subcategory(sub)
                sqlres_has_license_meter[rid.lower()] = True
            elif "compute" in sl or "vcore" in sl or "hyperscale" in sl:
                bucket = "compute"
                tier = _sqldb_tier_from_subcategory(sub)
                sqlres_has_compute_meter[rid.lower()] = True
            elif _is_dtu_tier(sub):
                # DTU-tier: bundles compute + license. AHB cannot apply.
                bucket = "compute"
                tier = "DTU (" + _sqldb_tier_from_subcategory(sub) + ")"
                # Track so we tag as 'Not eligible (DTU)' below.
                sqlres_has_compute_meter[rid.lower()] = True
            else:
                # Treat unrecognised SQL DB rows as 'other' (e.g. zone-redundancy
                # add-ons, secondary replicas already counted above, etc.).
                bucket = "other"
                tier = _sqldb_tier_from_subcategory(sub)

        # ---- Azure SQL Managed Instance ----
        elif cat == "SQL Managed Instance":
            workload = "Azure SQL Managed Instance"
            sl = sub.lower()
            if "storage" in sl or "backup" in sl:
                bucket = "storage"
                tier = _sqlmi_tier_from_subcategory(sub)
            elif "sql license" in sl:
                bucket = "license"
                tier = _sqlmi_tier_from_subcategory(sub)
                sqlres_has_license_meter[rid.lower()] = True
            elif "compute" in sl or "vcore" in sl:
                bucket = "compute"
                tier = _sqlmi_tier_from_subcategory(sub)
                sqlres_has_compute_meter[rid.lower()] = True
            else:
                bucket = "other"
                tier = _sqlmi_tier_from_subcategory(sub)

        # ---- Defender-for-SQL etc. — context cost on existing SQL resources ----
        elif cat == "Microsoft Defender for Cloud" and "SQL" in sub:
            # Only attribute if we recognise the resource as SQL DB / MI / VM.
            workload = ""  # decide based on existing aggregation
            tier = ""
            bucket = "other"
        else:
            continue

        if not bucket:
            continue

        qty  = _to_float(row.get(qty_col))  or 0.0
        cost = _to_float(row.get(cost_col)) or 0.0

        sub_id, rg, rtype, name = _parse_resource_id(rid)
        rid_l = rid.lower()
        agg = aggregations.get(rid_l)
        if agg is None:
            # If this row is Defender-only and we haven't seen the resource
            # via a SQL meter yet, skip — we don't want pure-Defender SQL VM
            # entries cluttering the report.
            if not workload:
                continue
            agg = SqlAggregation(
                resource_id=rid,
                name=name or "",
                resource_group=rg or "",
                subscription_id=sub_id or "",
                resource_type=rtype or "",
            )
            aggregations[rid_l] = agg

        if not agg.location:
            agg.location = row.get(reg_col) or ""
        if not agg.currency:
            agg.currency = row.get(cur_col) or ""
        if not agg.workload and workload:
            agg.workload = workload
        if not agg.tier and tier:
            agg.tier = tier
        if not agg.ahb_state and ahb_state:
            agg.ahb_state = ahb_state

        # Accumulate into the bucket.
        if bucket == "compute":
            agg.compute_cost += cost
        elif bucket == "storage":
            agg.storage_cost += cost
        elif bucket == "license":
            agg.license_cost += cost
        else:
            agg.other_cost += cost
        agg.total_cost     += cost
        agg.quantity_total += qty

    # Post-pass: derive AHB state.
    for rid_l, agg in aggregations.items():
        # SQL Server on VM: decide from observed cost-bearing license rows.
        if agg.workload == "SQL Server on VM":
            paid = vmsql_has_paid_lic.get(rid_l, False)
            ahb  = vmsql_has_ahb_row.get(rid_l, False)
            if paid and ahb:
                agg.ahb_state = "Partially applied"
                agg.notes.append("Both AHB-applied (0-cost) and paid-licence "
                                 "(Std/Ent) rows seen in the period — AHB "
                                 "likely toggled mid-period or split coverage.")
            elif paid:
                agg.ahb_state = "Not applied"
            elif ahb:
                agg.ahb_state = "Applied"
            else:
                agg.ahb_state = "Review (no SQL VM licence row)"
            # Tier: prefer Enterprise over Standard if both seen.
            eds = vmsql_paid_editions.get(rid_l, set())
            if "Enterprise" in eds:
                agg.tier = "Enterprise"
            elif "Standard" in eds:
                agg.tier = "Standard"
            elif ahb and not agg.tier:
                agg.tier = "AHB"
            continue

        if agg.ahb_state:  # already set (e.g. free editions)
            continue

        # Azure SQL DB / MI: infer from license-meter presence.
        has_license = sqlres_has_license_meter.get(rid_l, False)
        has_compute = sqlres_has_compute_meter.get(rid_l, False)
        is_dtu = agg.tier.startswith("DTU")
        if is_dtu:
            agg.ahb_state = "Not eligible (DTU tier)"
            agg.notes.append("DTU bundles compute+license. Migrate to vCore "
                             "tier to enable AHB.")
        elif has_license:
            agg.ahb_state = "Not applied"
        elif has_compute:
            agg.ahb_state = "Applied"
        else:
            # Storage-only row that we never matched a compute/license meter
            # to — most likely an orphan storage line (LTR backup of a
            # since-deleted DB). Mark as 'Review'.
            agg.ahb_state = "Review (storage-only meter rows)"

    period_info = _build_period_info(period_start_str, period_end_str,
                                     schema_name=schema["name"])
    elapsed = time.perf_counter() - t0
    print(f"Aggregated {len(aggregations)} SQL resources in {elapsed:.1f}s.")
    return (aggregations, period_info)


def _sqldb_tier_from_subcategory(sub: str) -> str:
    """Infer Azure SQL DB tier from FOCUS subcategory text."""
    s = sub.lower()
    if "hyperscale" in s:           return "Hyperscale"
    if "business critical" in s:    return "Business Critical"
    if "general purpose" in s and "serverless" in s: return "GP Serverless"
    if "general purpose" in s:      return "General Purpose"
    if "premium" in s:              return "Premium (DTU)"
    if "standard" in s:             return "Standard (DTU)"
    if "basic" in s:                return "Basic (DTU)"
    if "elastic pool" in s:         return "Elastic Pool"
    return ""


def _sqlmi_tier_from_subcategory(sub: str) -> str:
    s = sub.lower()
    if "business critical" in s:    return "Business Critical"
    if "general purpose" in s:      return "General Purpose"
    return ""


def _is_dtu_tier(sub: str) -> bool:
    """DTU-tier subcategory shapes (compute+license bundled, AHB ineligible)."""
    s = sub.lower()
    return ("single basic" in s
            or "single standard" in s
            or "single premium" in s
            or ("elastic pool" in s and ("standard" in s or "basic" in s
                                         or "premium" in s)))


# ---------- Build per-resource result rows ----------

def build_results(aggs: dict[str, SqlAggregation]) -> list[dict]:
    out: list[dict] = []
    for agg in aggs.values():
        # Savings = the license cost we observed on cost-bearing rows.
        # Applies to "Not applied" AND "Partially applied" (both have non-zero
        # license_cost from Std/Ent meters).
        savings_period = (agg.license_cost
                          if agg.ahb_state in ("Not applied", "Partially applied")
                          else 0.0)
        out.append({
            "Workload":             agg.workload,
            "Tier":                 agg.tier or None,
            "AhbState":             agg.ahb_state,
            "Name":                 agg.name,
            "ResourceGroup":        agg.resource_group,
            "SubscriptionId":       agg.subscription_id,
            "ResourceType":         agg.resource_type,
            "Location":             agg.location,
            "Currency":             agg.currency or None,
            "ComputeCostInPeriod":  round(agg.compute_cost, 4) if agg.compute_cost else None,
            "StorageCostInPeriod":  round(agg.storage_cost, 4) if agg.storage_cost else None,
            "LicenseCostInPeriod":  round(agg.license_cost, 4) if agg.license_cost else None,
            "OtherCostInPeriod":    round(agg.other_cost,   4) if agg.other_cost   else None,
            "TotalCostInPeriod":    round(agg.total_cost,   4) if agg.total_cost   else None,
            "PotentialSavingsPeriod": round(savings_period, 4) if savings_period else None,
            "Quantity":             round(agg.quantity_total, 4) if agg.quantity_total else None,
            "ResourceId":           agg.resource_id,
            "Notes":                "; ".join(agg.notes) or None,
        })
    return out


# ---------- main ----------

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--usage-csv", help="Detailed billing CSV.")
    parser.add_argument("--export-csv",  default="reports/sql-license-report.csv")
    parser.add_argument("--export-xlsx", default="reports/sql-license-report.xlsx")
    args = parser.parse_args(argv)

    script_path = Path(__file__).resolve()
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = find_default_usage_csv(script_path)
        if not found:
            print("ERROR: No usage CSV; pass --usage-csv.", file=sys.stderr)
            return 2
        usage_csv = found.resolve()
        print(f"Auto-detected usage CSV: {usage_csv}")
    if not usage_csv.exists():
        print(f"ERROR: Usage CSV not found: {usage_csv}", file=sys.stderr)
        return 2

    aggs, period = aggregate_sql(usage_csv)
    if not aggs:
        print("\nNo SQL workloads found in the CSV.")
        return 0

    results = build_results(aggs)
    period_days = period.get("days") or 30
    annualize   = 365.0 / period_days

    # ----- Console summary
    cur = next((r["Currency"] for r in results if r["Currency"]), "")
    by_state: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"count": 0, "license": 0.0, "compute": 0.0, "storage": 0.0,
                 "other": 0.0, "total": 0.0, "savings": 0.0})
    for r in results:
        k = (r["Workload"], r["AhbState"])
        b = by_state[k]
        b["count"]   += 1
        b["license"] += r["LicenseCostInPeriod"]    or 0
        b["compute"] += r["ComputeCostInPeriod"]    or 0
        b["storage"] += r["StorageCostInPeriod"]    or 0
        b["other"]   += r["OtherCostInPeriod"]      or 0
        b["total"]   += r["TotalCostInPeriod"]      or 0
        b["savings"] += r["PotentialSavingsPeriod"] or 0

    total_savings_period = sum(r["PotentialSavingsPeriod"] or 0 for r in results)
    total_savings_annual = total_savings_period * annualize
    total_license_period = sum(r["LicenseCostInPeriod"]    or 0 for r in results)
    total_total_period   = sum(r["TotalCostInPeriod"]      or 0 for r in results)

    print()
    print(f"Billing period: {period['label']}")
    print(f"SQL resources : {len(results)}")
    print(f"Total SQL spend in period      : {total_total_period:>14,.2f} {cur}")
    print(f"Of which is SQL license cost   : {total_license_period:>14,.2f} {cur}")
    print()
    print(f"{'Workload':<32} {'State':<28} {'#':>4} "
          f"{'License':>13} {'Compute':>13} {'PotentialSavings':>18}")
    print("-" * 116)
    for (wkl, state), b in sorted(by_state.items(),
                                  key=lambda kv: (-kv[1]["savings"], -kv[1]["license"])):
        print(f"{wkl[:32]:<32} {state[:28]:<28} {b['count']:>4} "
              f"{b['license']:>13,.2f} {b['compute']:>13,.2f} "
              f"{b['savings']:>18,.2f}")
    print("-" * 116)
    print(f"Total potential savings (period)   : {total_savings_period:>14,.2f} {cur}")
    print(f"Annualized potential savings       : {total_savings_annual:>14,.2f} {cur}/year")

    not_applied = [r for r in results
                   if (r["PotentialSavingsPeriod"] or 0) > 0]
    not_applied.sort(key=lambda r: -(r["PotentialSavingsPeriod"] or 0))
    if not_applied:
        print("\nTop SQL AHB opportunities (ranked by period license cost):")
        print(f"  {'Name':<40} {'Workload':<28} {'Tier':<18} "
              f"{'PotentialSavings':>18}")
        for r in not_applied[:15]:
            print(f"  {(r['Name'] or '')[:40]:<40} "
                  f"{(r['Workload'] or '')[:28]:<28} "
                  f"{(r['Tier'] or '')[:18]:<18} "
                  f"{(r['PotentialSavingsPeriod'] or 0):>18,.2f}")
        if len(not_applied) > 15:
            rest = sum(r["PotentialSavingsPeriod"] or 0 for r in not_applied[15:])
            print(f"  ... and {len(not_applied) - 15} more "
                  f"({rest:,.2f} {cur} period savings)")

    # ----- Sort + export
    state_order = {"Not applied": 0,
                   "Partially applied": 1,
                   "Not eligible (DTU tier)": 2,
                   "Review (storage-only meter rows)": 3,
                   "Review (no SQL VM licence row)": 3,
                   "Applied": 4,
                   "N/A (Free edition)": 5}
    results.sort(key=lambda r: (state_order.get(r["AhbState"], 99),
                                -(r["PotentialSavingsPeriod"] or 0),
                                -(r["TotalCostInPeriod"] or 0)))

    out_path = Path(args.export_csv).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)
    print(f"\nReport exported to {out_path}")

    if args.export_xlsx:
        xlsx_path = Path(args.export_xlsx).resolve()
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            write_xlsx(results, xlsx_path,
                       currency=cur, period=period, annualize=annualize,
                       by_state=by_state,
                       total_savings_period=total_savings_period,
                       total_savings_annual=total_savings_annual,
                       total_license_period=total_license_period,
                       total_total_period=total_total_period)
            print(f"Excel report exported to {xlsx_path}")
        except ImportError:
            print("WARNING: openpyxl not installed; skipping Excel export.",
                  file=sys.stderr)
        except PermissionError:
            print(f"WARNING: cannot write {xlsx_path} — is it open in Excel?",
                  file=sys.stderr)
    return 0


def write_xlsx(results, out_path, *, currency, period, annualize,
               by_state, total_savings_period, total_savings_annual,
               total_license_period, total_total_period) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Licenses"

    bold        = Font(bold=True)
    bold_white  = Font(bold=True, color="FFFFFF")
    big_bold    = Font(bold=True, size=14)
    accent      = Font(bold=True, color="0070C0")
    header_fill = PatternFill("solid", fgColor="305496")
    money_fmt   = "#,##0.00"

    period_label = period.get("label") or "unknown period"
    ws["A1"] = (f"SQL Server license scan — {len(results)} resources | "
                f"Period: {period_label}")
    ws["A1"].font = big_bold
    ws.merge_cells("A1:G1")

    # Headline numbers
    ws["A3"] = "Total SQL spend in period"
    ws["B3"] = round(total_total_period, 2); ws["C3"] = currency
    ws["A4"] = "Of which is SQL license cost"
    ws["B4"] = round(total_license_period, 2); ws["C4"] = currency
    ws["A5"] = "Potential savings if AHB enabled (period)"
    ws["B5"] = round(total_savings_period, 2); ws["C5"] = currency
    ws["A6"] = "Annualized potential savings"
    ws["B6"] = round(total_savings_annual, 2); ws["C6"] = f"{currency}/year"
    for r in (3, 4, 5, 6):
        ws.cell(row=r, column=2).number_format = money_fmt
    for r in (5, 6):
        ws.cell(row=r, column=1).font = accent
        ws.cell(row=r, column=2).font = accent

    # Summary by Workload × State
    sum_start = 8
    headers = ["Workload", "State", "Count",
               f"Compute ({currency})", f"Storage ({currency})",
               f"License ({currency})", f"Other ({currency})",
               f"Total ({currency})", f"Potential savings ({currency})"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=sum_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row = sum_start + 1
    for (wkl, state), b in sorted(by_state.items(),
                                  key=lambda kv: (-kv[1]["savings"], -kv[1]["license"])):
        ws.cell(row=row, column=1, value=wkl)
        ws.cell(row=row, column=2, value=state)
        ws.cell(row=row, column=3, value=b["count"])
        ws.cell(row=row, column=4, value=round(b["compute"], 2)).number_format = money_fmt
        ws.cell(row=row, column=5, value=round(b["storage"], 2)).number_format = money_fmt
        ws.cell(row=row, column=6, value=round(b["license"], 2)).number_format = money_fmt
        ws.cell(row=row, column=7, value=round(b["other"],   2)).number_format = money_fmt
        ws.cell(row=row, column=8, value=round(b["total"],   2)).number_format = money_fmt
        ws.cell(row=row, column=9, value=round(b["savings"], 2)).number_format = money_fmt
        row += 1

    table_start = row + 2

    # Per-resource table
    res_headers = list(results[0].keys())
    for c_idx, h in enumerate(res_headers, start=1):
        cell = ws.cell(row=table_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    money_cols = {"ComputeCostInPeriod", "StorageCostInPeriod",
                  "LicenseCostInPeriod", "OtherCostInPeriod",
                  "TotalCostInPeriod", "PotentialSavingsPeriod"}
    for r_idx, r in enumerate(results, start=table_start + 1):
        for c_idx, h in enumerate(res_headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=r.get(h))
            if h in money_cols and r.get(h) is not None:
                cell.number_format = money_fmt

    # Column widths
    for c_idx, h in enumerate(res_headers, start=1):
        letter = get_column_letter(c_idx)
        max_len = len(h)
        for r in results:
            v = r.get(h)
            if v is None: continue
            max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = f"A{table_start + 1}"
    if results:
        ws.auto_filter.ref = (
            f"A{table_start}:{get_column_letter(len(res_headers))}"
            f"{table_start + len(results)}")

    wb.save(out_path)


if __name__ == "__main__":
    sys.exit(main())
