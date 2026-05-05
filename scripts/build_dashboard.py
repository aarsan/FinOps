#!/usr/bin/env python3
"""Build a single Excel dashboard combining the Disk, VM, and AHB reports.

One workbook with:
  - Dashboard sheet (period banner, headline savings, charts)
  - Recommendations sheet (AHB + RI + SP guidance)
  - Per-domain detail sheets (Unattached disks, VMs, AHB resources, etc.)
  - Pie + bar charts driven by sheet ranges so they auto-update

Time grain is labelled explicitly throughout. Each money cell or chart axis
includes the period (e.g. "Nov 2025") or an explicit /month or /year suffix.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

# Make the existing module helpers importable.
sys.path.insert(0, str(Path(__file__).parent))

import list_unattached_disks as disks_mod
import list_vms as vms_mod
import list_ahb as ahb_mod
import list_sql as sql_mod

HOURS_PER_MONTH = 730


# ---------- Period helpers ----------

def short_period_label(period: dict) -> str:
    """Convert an ISO start (YYYY-MM-DD) into 'Nov 2025'-style. Falls back."""
    start = period.get("start") or ""
    if not start:
        return period.get("label") or "the period"
    try:
        from datetime import date
        d = date.fromisoformat(start)
        return d.strftime("%b %Y")
    except Exception:
        return period.get("label") or "the period"


def compute_total_period_cost(path: Path) -> tuple[float, float, dict[str, float],
                                                    dict[str, float], dict[str, float]]:
    """Single-pass scan over the whole CSV.

    Returns:
        total_cost      — sum of EffectiveCost / costInBillingCurrency
        total_list      — sum of ListCost / paygCostInBillingCurrency
        by_service      — {ServiceName: cost}
        realized_savings — {mechanism: per-row (list - effective) attributed
                            to that mechanism}. For Reservations/Savings
                            Plans, FOCUS records the consumption row with
                            ListCost=0, so per-row savings on commitment
                            rows is usually $0. The bulk that lands in
                            'Negotiated discount (no commitment)' is real:
                            MCA/contract savings on regular usage.
        cost_covered    — {mechanism: sum of EffectiveCost on rows tagged
                            with that mechanism}. Useful for showing 'X% of
                            spend is RI/SP-covered'.
    """
    import csv as _csv
    import time as _time

    print(f"\nScanning '{path.name}' for total spend (one pass)...")
    t0 = _time.perf_counter()

    total_cost = 0.0
    total_list = 0.0
    by_service: dict[str, float] = defaultdict(float)
    realized: dict[str, float] = defaultdict(float)
    cost_covered: dict[str, float] = defaultdict(float)

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = _csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return (0.0, 0.0, {}, {})

        # FOCUS first, legacy fallback.
        col_cost = (header.index("EffectiveCost")
                    if "EffectiveCost" in header
                    else header.index("costInBillingCurrency")
                    if "costInBillingCurrency" in header else None)
        col_list = (header.index("ListCost")
                    if "ListCost" in header
                    else header.index("paygCostInBillingCurrency")
                    if "paygCostInBillingCurrency" in header else None)
        col_svc = (header.index("ServiceName")
                   if "ServiceName" in header
                   else header.index("meterCategory")
                   if "meterCategory" in header else None)
        # Mechanism attribution
        col_pricing_cat = (header.index("PricingCategory")
                           if "PricingCategory" in header
                           else header.index("pricingModel")
                           if "pricingModel" in header else None)
        col_benefit_type = (header.index("CommitmentDiscountType")
                            if "CommitmentDiscountType" in header else None)

        if col_cost is None:
            raise RuntimeError("Cost column not found in CSV header.")

        rows = 0
        progress_every = 500_000
        for row in reader:
            rows += 1
            try:
                cost = float(row[col_cost]) if row[col_cost] else 0.0
                total_cost += cost
                lst = 0.0
                if col_list is not None and row[col_list]:
                    lst = float(row[col_list])
                    total_list += lst
                if col_svc is not None:
                    by_service[row[col_svc] or "Unknown"] += cost

                # ----- Realized savings attribution -----
                # FOCUS records reservations and savings plans in two ways:
                #   1. Consumption rows  (ChargeCategory='Usage', CommitmentDiscountType
                #      set, EffectiveCost = amortized rate, ListCost = 0). These
                #      have NO per-row savings — list − effective is meaningless.
                #   2. Purchase / amortization rows where Cost > 0 covers the
                #      amortized commitment.
                # The CORRECT way to measure realized RI/SP savings is:
                #   (a) sum EffectiveCost on rows tagged as covered by a
                #       commitment (PricingCategory == 'Committed' OR
                #       CommitmentDiscountType set);
                #   (b) compare with what those hours would have cost at list,
                #       which we approximate from the customer's other usage at
                #       the same SKU. Doing that here would be expensive.
                # As a defensible proxy: attribute the per-row (list - effective)
                # gap to negotiated/MCA savings (it really is that) and
                # attribute commitment-covered EFFECTIVE cost to the commitment
                # mechanism so the dashboard at least shows what's RI/SP-covered.
                pricing = (row[col_pricing_cat] or "").strip() \
                    if col_pricing_cat is not None else ""
                btype   = (row[col_benefit_type] or "").strip() \
                    if col_benefit_type is not None else ""
                # Classify the row's benefit mechanism (if any)
                if btype == "Reservation" or pricing == "Reservation":
                    mech = "Reservation"
                elif btype == "SavingsPlan" or pricing == "SavingsPlan":
                    mech = "Savings Plan"
                elif pricing.lower() in ("spot", "lowpriority"):
                    mech = "Spot"
                elif pricing == "Committed":
                    # Committed rate but type not specified — most likely RI.
                    mech = "Reservation"
                else:
                    mech = ""

                # Track effective cost covered by each mechanism (used to
                # derive 'cost going through commitments' for the dashboard).
                if mech:
                    cost_covered[mech] += cost

                # Per-row realized savings on rows where ListCost is populated.
                # This captures MCA/contract negotiated savings cleanly for
                # non-commitment rows, and also a tiny portion of commitment
                # rows where FOCUS happens to populate both.
                if lst > 0 and cost < lst - 1e-9:
                    saving = lst - cost
                    if mech:
                        # Row is covered by a commitment AND has a per-row gap
                        # — attribute to the mechanism (rare but legitimate).
                        realized[mech] += saving
                    else:
                        realized["Negotiated discount (no commitment)"] += saving
            except (IndexError, ValueError):
                continue
            if rows % progress_every == 0:
                print(f"  scanned {rows:,} rows ({_time.perf_counter() - t0:.1f}s)...")

    elapsed = _time.perf_counter() - t0
    realized_total = sum(realized.values())
    covered_total  = sum(cost_covered.values())
    print(f"  total spend pass complete: {rows:,} rows in {elapsed:.1f}s.")
    print(f"    Total spend             : {total_cost:,.2f}")
    print(f"    Total at list           : {total_list:,.2f}")
    print(f"    Realized per-row savings: {realized_total:,.2f}")
    print(f"    Cost covered by RI/SP   : {covered_total:,.2f}")
    return (total_cost, total_list, dict(by_service), dict(realized), dict(cost_covered))



# ---------- Main ----------

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--usage-csv",
                        help="Detailed billing CSV. Defaults to the most recent "
                             "*.csv under <workspace>/data/.")
    parser.add_argument("--export-xlsx",
                        default="reports/finops-dashboard.xlsx",
                        help="Output dashboard workbook (default: %(default)s).")
    parser.add_argument("--skip-disks", action="store_true",
                        help="Skip the disk report (no Resource Graph call).")
    parser.add_argument("--all-disks", action="store_true",
                        help="Disk report includes any state, not just Unattached.")
    parser.add_argument("--tenant-id",
                        help="Azure AD tenant ID for the disk Resource Graph call.")
    args = parser.parse_args(argv)

    script_path = Path(__file__).resolve()

    # ----- Resolve usage CSV
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = disks_mod.find_default_usage_csv(script_path)
        if not found:
            print("ERROR: No usage CSV; pass --usage-csv.", file=sys.stderr)
            return 2
        usage_csv = found.resolve()
        print(f"Auto-detected usage CSV: {usage_csv}")
    if not usage_csv.exists():
        print(f"ERROR: Usage CSV not found: {usage_csv}", file=sys.stderr)
        return 2

    # ============================================================
    # Stage 1: VM aggregation (also gives us the period info)
    # ============================================================
    print("\n" + "=" * 70)
    print("  VM aggregation")
    print("=" * 70)
    vm_billing, period = vms_mod.aggregate_vm_billing(usage_csv)
    vm_rows = vms_mod.build_results(vm_billing)
    vm_rows.sort(key=lambda r: -(r["ActualCostInPeriod"] or 0))
    period_days = period.get("days") or 30
    annualize   = 365.0 / period_days
    period_label = period.get("label") or "unknown period"
    period_short = short_period_label(period)
    print(f"  Period: {period_label}")

    # VM insights
    vm_by_category    = vms_mod.compute_vm_by_category(vm_rows)
    observed_ri_disc  = vms_mod.compute_observed_ri_discount(vm_rows)
    ri_top            = vms_mod.compute_ri_candidates(vm_rows, observed_ri_disc, annualize)
    vm_ahb_candidates = vms_mod.compute_ahb_vm_candidates(vm_rows)
    spot_candidates   = vms_mod.compute_spot_candidates(vm_rows)

    # ============================================================
    # Stage 1b: Total spend (one fast pass over the whole CSV)
    # ============================================================
    total_period_cost, total_period_list, by_service, realized_savings, cost_covered = \
        compute_total_period_cost(usage_csv)

    # ============================================================
    # Stage 2: AHB scan (uses the same CSV, adds SQL coverage)
    # ============================================================
    print("\n" + "=" * 70)
    print("  AHB scan")
    print("=" * 70)
    ahb_aggs, _ = ahb_mod.aggregate_ahb(usage_csv)
    ahb_rows = ahb_mod.build_results(ahb_aggs)

    # AHB insights
    ahb_by_state: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"count": 0, "license": 0.0, "compute": 0.0, "total": 0.0,
                 "savings": 0.0})
    for r in ahb_rows:
        k = (r["Workload"], r["AhbState"])
        b = ahb_by_state[k]
        b["count"]   += 1
        b["license"] += r["LicenseCostInPeriod"]    or 0
        b["compute"] += r["ComputeCostInPeriod"]    or 0
        b["total"]   += r["TotalCostInPeriod"]      or 0
        b["savings"] += r["PotentialSavingsPeriod"] or 0
    ahb_top = sorted(
        [r for r in ahb_rows if (r["PotentialSavingsPeriod"] or 0) > 0],
        key=lambda r: -(r["PotentialSavingsPeriod"] or 0)
    )

    # ============================================================
    # Stage 2b: SQL Server license scan (deep view of SQL costs)
    # The AHB scan already counts SQL-Server-on-VM license cost.
    # SQL scan extends that to Azure SQL DB / MI license meters which
    # AHB doesn't quantify. We compute a 'SQL DB+MI only' delta below to
    # avoid double-counting SQL on VM in the headline opportunity.
    # ============================================================
    print("\n" + "=" * 70)
    print("  SQL license scan")
    print("=" * 70)
    sql_aggs, _ = sql_mod.aggregate_sql(usage_csv)
    sql_rows = sql_mod.build_results(sql_aggs)

    sql_by_state: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"count": 0, "license": 0.0, "compute": 0.0,
                 "storage": 0.0, "other": 0.0, "total": 0.0,
                 "savings": 0.0})
    for r in sql_rows:
        k = (r["Workload"], r["AhbState"])
        b = sql_by_state[k]
        b["count"]   += 1
        b["license"] += r["LicenseCostInPeriod"]    or 0
        b["compute"] += r["ComputeCostInPeriod"]    or 0
        b["storage"] += r["StorageCostInPeriod"]    or 0
        b["other"]   += r["OtherCostInPeriod"]      or 0
        b["total"]   += r["TotalCostInPeriod"]      or 0
        b["savings"] += r["PotentialSavingsPeriod"] or 0
    sql_top = sorted(
        [r for r in sql_rows if (r["PotentialSavingsPeriod"] or 0) > 0],
        key=lambda r: -(r["PotentialSavingsPeriod"] or 0)
    )

    # ============================================================
    # Stage 3: Disk report (Resource Graph + price index)
    # ============================================================
    disk_rows: list[dict] = []
    disk_query_error: Optional[str] = None
    if not args.skip_disks:
        print("\n" + "=" * 70)
        print("  Disk inventory + pricing")
        print("=" * 70)
        try:
            azure_disks = disks_mod.query_disks_via_resource_graph(
                all_states=args.all_disks, tenant_id=args.tenant_id)
        except Exception as exc:  # noqa: BLE001
            disk_query_error = str(exc)
            print(f"  WARNING: Resource Graph call failed: {exc}",
                  file=sys.stderr)
            azure_disks = []
        if azure_disks:
            price_index = disks_mod.build_price_index(usage_csv)
            disk_rows, no_price = disks_mod.build_disk_rows(azure_disks, price_index)
            disk_rows.sort(key=lambda r: -(r["CustomerMonthlyCost"] or 0))
            if no_price:
                print(f"  Note: {no_price} disk(s) had no matching CSV price.")
    else:
        print("\nSkipping disk inventory (--skip-disks).")

    # ============================================================
    # Stage 4: Pick currency and headline numbers
    # ============================================================
    cur = (
        next((r["Currency"] for r in vm_rows  if r.get("Currency")), "")
        or next((r["Currency"] for r in disk_rows if r.get("Currency")), "")
        or next((r["Currency"] for r in ahb_rows  if r.get("Currency")), "")
        or next((r["Currency"] for r in sql_rows  if r.get("Currency")), "")
    )

    # ============================================================
    # Stage 5: Write the workbook
    # ============================================================
    out_path = Path(args.export_xlsx).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        write_dashboard(
            out_path=out_path,
            currency=cur,
            period=period, period_label=period_label, period_short=period_short,
            annualize=annualize,
            total_period_cost=total_period_cost,
            total_period_list=total_period_list,
            by_service=by_service,
            realized_savings=realized_savings,
            cost_covered=cost_covered,
            vm_rows=vm_rows, vm_by_category=vm_by_category,
            observed_ri_discount=observed_ri_disc,
            ri_top=ri_top, vm_ahb_candidates=vm_ahb_candidates,
            spot_candidates=spot_candidates,
            ahb_rows=ahb_rows, ahb_by_state=ahb_by_state, ahb_top=ahb_top,
            sql_rows=sql_rows, sql_by_state=sql_by_state, sql_top=sql_top,
            disk_rows=disk_rows, disk_query_error=disk_query_error,
        )
    except PermissionError:
        print(f"\nERROR: cannot write {out_path} — is it open in Excel?",
              file=sys.stderr)
        return 1
    print(f"\nDashboard exported to {out_path}")
    return 0


# ============================================================
# Workbook builder
# ============================================================

def write_dashboard(
    *,
    out_path: Path, currency: str,
    period: dict, period_label: str, period_short: str,
    annualize: float,
    total_period_cost: float, total_period_list: float,
    by_service: dict[str, float],
    realized_savings: dict[str, float],
    cost_covered: dict[str, float],
    vm_rows: list[dict], vm_by_category: dict[str, tuple[int, float]],
    observed_ri_discount: Optional[float],
    ri_top: list[dict], vm_ahb_candidates: list[dict],
    spot_candidates: list[dict],
    ahb_rows: list[dict], ahb_by_state: dict, ahb_top: list[dict],
    sql_rows: list[dict], sql_by_state: dict, sql_top: list[dict],
    disk_rows: list[dict], disk_query_error: Optional[str],
) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # We'll insert the Dashboard at index 0 last; for now write data sheets.
    default = wb.active
    wb.remove(default)

    # ---- Style helpers
    bold        = Font(bold=True)
    bold_white  = Font(bold=True, color="FFFFFF")
    big_bold    = Font(bold=True, size=14)
    huge_bold   = Font(bold=True, size=18)
    accent      = Font(bold=True, color="0070C0")
    section_fill = PatternFill("solid", fgColor="305496")
    money_fmt   = "#,##0.00"
    pct_fmt     = '0.00"%"'
    rate_fmt    = "#,##0.000000"

    money_money_fmt_per = '"' + (currency + " ") + '"#,##0.00'

    def set_header(ws, row, cols):
        for c_idx, val in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font = bold_white
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def autosize(ws, headers, rows):
        for c_idx, h in enumerate(headers, start=1):
            letter = get_column_letter(c_idx)
            max_len = len(str(h))
            for r in rows:
                v = r.get(h) if isinstance(r, dict) else None
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)), 50))
            ws.column_dimensions[letter].width = max_len + 2

    # =========================================================
    # Sheet: VMs (full table)
    # =========================================================
    vm_ws = wb.create_sheet("VMs")
    vm_ws["A1"] = (f"VMs — {len(vm_rows)} VMs billed in {period_label}. "
                   f"All money columns are for this period unless suffixed.")
    vm_ws["A1"].font = big_bold
    vm_ws.merge_cells("A1:H1")
    if vm_rows:
        headers = list(vm_rows[0].keys())
        set_header(vm_ws, 3, headers)
        money_cols = {"ActualCostInPeriod", "ActualListInPeriod",
                      "ActualSavings", "WindowsSurchargeCost",
                      "EstMonthlyCost", "EstMonthlyListCost", "EstMonthlySavings"}
        rate_cols  = {"EffectiveHourlyRate", "ContractedHourlyRate", "PayGHourlyRate"}
        pct_cols   = {"DiscountPercent"}
        for r_idx, r in enumerate(vm_rows, start=4):
            for c_idx, h in enumerate(headers, start=1):
                cell = vm_ws.cell(row=r_idx, column=c_idx, value=r.get(h))
                if h in money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
                elif h in rate_cols and r.get(h) is not None:
                    cell.number_format = rate_fmt
                elif h in pct_cols and r.get(h) is not None:
                    cell.number_format = pct_fmt
        autosize(vm_ws, headers, vm_rows)
        vm_ws.freeze_panes = "A4"
        vm_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(vm_rows)}"

    # =========================================================
    # Sheet: Unattached Disks
    # =========================================================
    disk_ws = wb.create_sheet("Unattached Disks")
    if disk_query_error:
        disk_ws["A1"] = (f"Unattached disks — Resource Graph query failed: "
                         f"{disk_query_error}")
        disk_ws["A1"].font = big_bold
    elif not disk_rows:
        disk_ws["A1"] = (f"Unattached disks — none currently in Unattached state, "
                         f"or skipped via --skip-disks.")
        disk_ws["A1"].font = big_bold
    else:
        disk_ws["A1"] = (f"Unattached disks — {len(disk_rows)} disks. "
                         f"Money columns are MONTHLY rates (PAYG meter is 1/Month). "
                         f"Pricing source: {period_label}.")
        disk_ws["A1"].font = big_bold
        disk_ws.merge_cells("A1:H1")
        headers = list(disk_rows[0].keys())
        set_header(disk_ws, 3, headers)
        money_cols = {"EffectivePrice", "ContractedPrice", "PayGPrice",
                      "CustomerMonthlyCost", "ListMonthlyCost", "MonthlySavings"}
        pct_cols   = {"BenefitDiscountPct", "ContractDiscountPct"}
        for r_idx, r in enumerate(disk_rows, start=4):
            for c_idx, h in enumerate(headers, start=1):
                cell = disk_ws.cell(row=r_idx, column=c_idx, value=r.get(h))
                if h in money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
                elif h in pct_cols and r.get(h) is not None:
                    cell.number_format = pct_fmt
        autosize(disk_ws, headers, disk_rows)
        disk_ws.freeze_panes = "A4"
        disk_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(disk_rows)}"

    # =========================================================
    # Sheet: AHB resources (full table)
    # =========================================================
    ahb_ws = wb.create_sheet("AHB Resources")
    ahb_ws["A1"] = (f"AHB scan — {len(ahb_rows)} resources. "
                    f"Money columns are for {period_label}.")
    ahb_ws["A1"].font = big_bold
    ahb_ws.merge_cells("A1:H1")
    if ahb_rows:
        headers = list(ahb_rows[0].keys())
        set_header(ahb_ws, 3, headers)
        money_cols = {"ComputeCostInPeriod", "LicenseCostInPeriod",
                      "TotalCostInPeriod", "PotentialSavingsPeriod"}
        for r_idx, r in enumerate(ahb_rows, start=4):
            for c_idx, h in enumerate(headers, start=1):
                cell = ahb_ws.cell(row=r_idx, column=c_idx, value=r.get(h))
                if h in money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
        autosize(ahb_ws, headers, ahb_rows)
        ahb_ws.freeze_panes = "A4"
        ahb_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(ahb_rows)}"

    # =========================================================
    # Sheet: SQL Licenses (full table — every SQL resource)
    # =========================================================
    sql_ws = wb.create_sheet("SQL Licenses")
    sql_ws["A1"] = (f"SQL Server license scan — {len(sql_rows)} SQL resources "
                    f"(SQL on VM + Azure SQL DB + Managed Instance). "
                    f"Money columns are for {period_label}.")
    sql_ws["A1"].font = big_bold
    sql_ws.merge_cells("A1:H1")
    if sql_rows:
        headers = list(sql_rows[0].keys())
        set_header(sql_ws, 3, headers)
        money_cols = {"ComputeCostInPeriod", "StorageCostInPeriod",
                      "LicenseCostInPeriod", "OtherCostInPeriod",
                      "TotalCostInPeriod", "PotentialSavingsPeriod"}
        for r_idx, r in enumerate(sql_rows, start=4):
            for c_idx, h in enumerate(headers, start=1):
                cell = sql_ws.cell(row=r_idx, column=c_idx, value=r.get(h))
                if h in money_cols and r.get(h) is not None:
                    cell.number_format = money_fmt
        autosize(sql_ws, headers, sql_rows)
        sql_ws.freeze_panes = "A4"
        sql_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(sql_rows)}"

    # =========================================================
    # Sheet: Charts data (the small aggregated tables charts read from)
    # =========================================================
    cd = wb.create_sheet("ChartsData")
    cd["A1"] = "Aggregated tables used by dashboard charts"
    cd["A1"].font = big_bold

    # Block 1: VM cost by benefit category (for pie)
    row = 3
    cd.cell(row=row, column=1, value="VM cost by benefit category").font = bold
    row += 1
    set_header(cd, row, ["Category", "VM count",
                         f"Cost in {period_short} ({currency})"])
    cat_start = row + 1
    for i, (cat, (n, c)) in enumerate(
            sorted(vm_by_category.items(), key=lambda kv: -kv[1][1]),
            start=cat_start):
        cd.cell(row=i, column=1, value=cat)
        cd.cell(row=i, column=2, value=n)
        cd.cell(row=i, column=3, value=round(c, 2)).number_format = money_fmt
    cat_end = cat_start + len(vm_by_category) - 1 if vm_by_category else cat_start - 1

    # Block 2: AHB summary by Workload + State (used by pie of license cost).
    # The pie is cost-weighted so the "Applied or N/A (Linux or AHB)" bucket
    # — which has $0 license cost — naturally drops out, leaving only the
    # actionable spend. Rows are sorted with non-zero license-cost first so
    # the pie's range reads from the top.
    row = cat_end + 3
    cd.cell(row=row, column=1, value="AHB by workload × state").font = bold
    row += 1
    set_header(cd, row, ["Workload — State (short)", "Count",
                         f"License cost in {period_short} ({currency})",
                         f"Potential savings in {period_short} ({currency})",
                         f"Annualized potential savings ({currency}/yr)"])
    ahb_start = row + 1
    # Short label: e.g. "SQL Server on VM — Not applied"
    ahb_block = sorted(ahb_by_state.items(),
                       key=lambda kv: (-kv[1]["license"], -kv[1]["count"]))
    for i, ((wkl, state), b) in enumerate(ahb_block, start=ahb_start):
        # Trim known noise tails so labels fit nicely on chart.
        short_state = (state.replace("(License-included tier)", "(License-incl)")
                            .replace("(vCore + license split varies)", "(needs review)")
                            .replace("(BYOS / license-mobility)", "(BYOS)")
                            .replace("Applied or N/A (Linux or AHB)", "Applied / Linux"))
        cd.cell(row=i, column=1, value=f"{wkl} — {short_state}")
        cd.cell(row=i, column=2, value=b["count"])
        cd.cell(row=i, column=3, value=round(b["license"], 2)).number_format = money_fmt
        cd.cell(row=i, column=4, value=round(b["savings"], 2)).number_format = money_fmt
        cd.cell(row=i, column=5, value=round(b["savings"] * annualize, 2)).number_format = money_fmt
    ahb_end = ahb_start + len(ahb_block) - 1 if ahb_block else ahb_start - 1
    # Track the subset that has non-zero license cost — that's the pie range.
    ahb_pie_rows = [i for i, (_, b) in enumerate(ahb_block, start=ahb_start)
                    if b["license"] > 0]
    ahb_pie_first = ahb_pie_rows[0] if ahb_pie_rows else None
    ahb_pie_last  = ahb_pie_rows[-1] if ahb_pie_rows else None

    # Block 3: Top RI candidates (for bar)
    row = ahb_end + 3
    cd.cell(row=row, column=1, value="Top RI candidates (annualized savings)").font = bold
    row += 1
    set_header(cd, row, ["VmSize × Region", "VMs",
                         f"Cost in {period_short} ({currency})",
                         f"Projected savings in {period_short} ({currency})",
                         f"Projected annual savings ({currency}/yr)"])
    ri_start = row + 1
    ri_top10 = ri_top[:10]
    for i, x in enumerate(ri_top10, start=ri_start):
        cd.cell(row=i, column=1, value=f"{x['VmSize']} × {x['Location']}")
        cd.cell(row=i, column=2, value=x["VmCount"])
        cd.cell(row=i, column=3, value=round(x["ActualCost"], 2)).number_format = money_fmt
        cd.cell(row=i, column=4, value=(x["ProjectedSavingsPeriod"] or 0)).number_format = money_fmt
        cd.cell(row=i, column=5, value=(x["ProjectedSavingsAnnual"] or 0)).number_format = money_fmt
    ri_end = ri_start + len(ri_top10) - 1 if ri_top10 else ri_start - 1

    # Block 4: Top AHB opportunities (for bar)
    row = ri_end + 3
    cd.cell(row=row, column=1, value="Top AHB opportunities (annualized savings)").font = bold
    row += 1
    set_header(cd, row, ["Resource", "Workload", "Edition",
                         f"License cost in {period_short} ({currency})",
                         f"Annualized savings ({currency}/yr)"])
    ahb_top_start = row + 1
    ahb_top10 = ahb_top[:10]
    for i, r in enumerate(ahb_top10, start=ahb_top_start):
        cd.cell(row=i, column=1, value=r["Name"])
        cd.cell(row=i, column=2, value=r["Workload"])
        cd.cell(row=i, column=3, value=r.get("Edition") or "")
        cd.cell(row=i, column=4, value=(r["LicenseCostInPeriod"] or 0)).number_format = money_fmt
        cd.cell(row=i, column=5, value=(r["PotentialSavingsPeriod"] or 0) * annualize).number_format = money_fmt
    ahb_top_end = ahb_top_start + len(ahb_top10) - 1 if ahb_top10 else ahb_top_start - 1

    # Block 5: Top unattached disks by monthly cost (for bar)
    row = ahb_top_end + 3
    cd.cell(row=row, column=1, value="Top unattached disks by monthly cost").font = bold
    row += 1
    set_header(cd, row, ["Disk", "Tier",
                         f"Monthly cost ({currency}/mo)",
                         f"Annualized cost ({currency}/yr)"])
    dk_start = row + 1
    dk_top10 = disk_rows[:10] if disk_rows else []
    for i, r in enumerate(dk_top10, start=dk_start):
        cd.cell(row=i, column=1, value=r["Name"])
        cd.cell(row=i, column=2, value=r.get("Tier") or "")
        cd.cell(row=i, column=3, value=(r["CustomerMonthlyCost"] or 0)).number_format = money_fmt
        cd.cell(row=i, column=4, value=(r["CustomerMonthlyCost"] or 0) * 12).number_format = money_fmt
    dk_end = dk_start + len(dk_top10) - 1 if dk_top10 else dk_start - 1

    # Block 6: Total spend by service (top 12 — for pie of overall spend)
    row = dk_end + 3
    cd.cell(row=row, column=1, value="Total spend by service").font = bold
    row += 1
    set_header(cd, row, ["Service", f"Cost in {period_short} ({currency})"])
    sv_start = row + 1
    if by_service:
        top_services = sorted(by_service.items(), key=lambda kv: -kv[1])[:12]
        # Roll the long tail into "Other" so the pie isn't cluttered.
        if len(by_service) > 12:
            other_total = sum(v for k, v in by_service.items()
                              if k not in {k0 for k0, _ in top_services})
            top_services.append(("Other (smaller services)", other_total))
        for i, (svc, cost) in enumerate(top_services, start=sv_start):
            cd.cell(row=i, column=1, value=svc)
            cd.cell(row=i, column=2, value=round(cost, 2)).number_format = money_fmt
        sv_end = sv_start + len(top_services) - 1
    else:
        sv_end = sv_start - 1

    # Block 7: Savings — realized vs additional opportunity (for bar chart).
    # Realized RI/SP savings are estimated using the customer's observed
    # RI discount factor (FOCUS records commitment-covered usage with
    # ListCost=0, so per-row arithmetic doesn't work for those).
    if observed_ri_discount and 0 < observed_ri_discount < 1:
        ri_factor = observed_ri_discount / (1 - observed_ri_discount)
    else:
        ri_factor = 0.0
    row = sv_end + 3
    cd.cell(row=row, column=1, value="Savings: realized vs additional opportunity").font = bold
    row += 1
    set_header(cd, row, ["Mechanism",
                         f"Realized in {period_short} ({currency})",
                         f"Additional in {period_short} ({currency})"])
    sav_start = row + 1
    sav_data = [
        ("Reservations (estimated*)",
            cost_covered.get("Reservation", 0.0)  * ri_factor,                  0.0),
        ("Savings Plans (estimated*)",
            cost_covered.get("Savings Plan", 0.0) * ri_factor,                  0.0),
        ("Spot (cost covered, savings n/a)",
            cost_covered.get("Spot", 0.0),                                      0.0),
        ("Negotiated discount (MCA / partner-led)",
            realized_savings.get("Negotiated discount (no commitment)", 0.0),   0.0),
        # Opportunities (only-additional rows)
        ("AHB (additional)",
            0.0, sum(r["PotentialSavingsPeriod"] or 0 for r in ahb_rows)),
        ("RI (additional, on-demand VMs)",
            0.0, sum(c.get("ProjectedSavingsPeriod") or 0 for c in ri_top)),
    ]
    for i, (mech, realized, opp) in enumerate(sav_data, start=sav_start):
        cd.cell(row=i, column=1, value=mech)
        cd.cell(row=i, column=2, value=round(realized, 2)).number_format = money_fmt
        cd.cell(row=i, column=3, value=round(opp, 2)).number_format = money_fmt
    sav_end = sav_start + len(sav_data) - 1

    # Make the columns roomy
    for c, w in zip("ABCDEF", (45, 22, 22, 22, 22, 22)):
        cd.column_dimensions[c].width = w

    # =========================================================
    # Sheet: Recommendations (text + actions)
    # =========================================================
    rec = wb.create_sheet("Recommendations")
    rec["A1"] = f"Optimization recommendations  |  Period: {period_label}"
    rec["A1"].font = huge_bold
    rec.merge_cells("A1:E1")

    rec["A3"] = (f"Time grain: '(period)' figures are for {period_label} "
                 f"({period_days_label(period)}). 'Monthly' = full-month run "
                 f"rate (rate × {HOURS_PER_MONTH}h). 'Annual' = period × 365/days.")
    rec["A3"].font = Font(italic=True, color="555555")
    rec.merge_cells("A3:E3")

    row = 5

    # 1. AHB
    rec.cell(row=row, column=1, value="1. Azure Hybrid Benefit (AHB)").font = Font(bold=True, size=12)
    row += 1
    ahb_savings_period = sum(r["PotentialSavingsPeriod"] or 0 for r in ahb_rows)
    ahb_savings_annual = ahb_savings_period * annualize
    not_applied = sum(1 for r in ahb_rows if r["AhbState"].startswith("Not applied"))
    if not_applied:
        rec.cell(row=row, column=1,
                 value=f"  • {not_applied} resource(s) currently NOT using AHB.")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • Surcharge billed in {period_short}: "
                       f"{ahb_savings_period:,.2f} {currency}.")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • Annualized savings if AHB applied: "
                       f"{ahb_savings_annual:,.2f} {currency}/year.")
        rec.cell(row=row, column=1).font = accent
        row += 1
        rec.cell(row=row, column=1,
                 value="  • Action: enable AHB on Windows VMs (with Software "
                       "Assurance) and SQL Server VMs (with SQL CALs). For "
                       "Azure SQL DB / MI, switch from license-included to "
                       "vCore + AHB if eligible.")
        row += 1
    else:
        rec.cell(row=row, column=1,
                 value="  • No resources flagged as 'Not applied'. AHB is "
                       "either applied or N/A across the eligible workloads.")
        row += 1
    row += 1

    # 2. RI
    rec.cell(row=row, column=1, value="2. Reservations (RI)").font = Font(bold=True, size=12)
    row += 1
    if observed_ri_discount is not None:
        rec.cell(row=row, column=1,
                 value=f"  • Observed customer RI discount (derived from "
                       f"existing RI-covered VMs in {period_short}): "
                       f"{observed_ri_discount * 100:.1f}% off list.")
        row += 1
    else:
        rec.cell(row=row, column=1,
                 value="  • No existing RI-covered VMs in the period; cannot "
                       "derive a customer-specific RI discount rate.")
        row += 1
    if ri_top:
        always_on = [c for c in ri_top if c["Coverage"] == "Always-on"]
        ri_period_total = sum(c.get("ProjectedSavingsPeriod") or 0 for c in ri_top)
        ri_annual_total = sum(c.get("ProjectedSavingsAnnual") or 0 for c in ri_top)
        rec.cell(row=row, column=1,
                 value=f"  • {len(ri_top)} (VmSize × region) groups currently "
                       f"on-demand could benefit from RIs ({len(always_on)} are "
                       f"always-on).")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • Projected savings in {period_short}: "
                       f"{ri_period_total:,.2f} {currency}.")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • Projected ANNUAL savings: "
                       f"{ri_annual_total:,.2f} {currency}/year.")
        rec.cell(row=row, column=1).font = accent
        row += 1
        rec.cell(row=row, column=1,
                 value="  • Action: buy 1- or 3-year RIs for stable always-on "
                       "workloads (top groups on the Dashboard / RI Candidates "
                       "sheet).")
        row += 1
    else:
        rec.cell(row=row, column=1,
                 value="  • No on-demand VMs that would clearly benefit from "
                       "additional RI coverage based on current run hours.")
        row += 1
    row += 1

    # 3. SP
    rec.cell(row=row, column=1, value="3. Compute Savings Plans (SP)").font = Font(bold=True, size=12)
    row += 1
    rec.cell(row=row, column=1, value=(
        "  • SPs are flatter, more flexible than RIs: they cover ANY VM "
        "family/region (and instance scaling) for a 1- or 3-yr hourly commit."
    )).alignment = Alignment(wrap_text=True)
    row += 1
    rec.cell(row=row, column=1, value=(
        "  • Use SPs when workload mix changes often (resizing, family swaps, "
        "region migrations); use RIs for stable size+region for the term."
    )).alignment = Alignment(wrap_text=True)
    row += 1
    rec.cell(row=row, column=1, value=(
        "  • Public reference: compute SP ≈ 17% off list (1-yr) / 33% (3-yr); "
        "RIs typically deliver more for stable workloads."
    )).alignment = Alignment(wrap_text=True)
    row += 1

    # 4. Unattached disks
    if disk_rows:
        row += 1
        rec.cell(row=row, column=1, value="4. Unattached disks").font = Font(bold=True, size=12)
        row += 1
        total_dk_monthly = sum(r["CustomerMonthlyCost"] or 0 for r in disk_rows)
        rec.cell(row=row, column=1,
                 value=f"  • {len(disk_rows)} disk(s) currently in Unattached "
                       f"state — paid storage with no compute attached.")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • Monthly cost (savings if all deleted): "
                       f"{total_dk_monthly:,.2f} {currency}/month "
                       f"(~{total_dk_monthly * 12:,.2f} {currency}/year).")
        rec.cell(row=row, column=1).font = accent
        row += 1
        rec.cell(row=row, column=1,
                 value="  • Action: snapshot any with valuable data, then "
                       "delete. See 'Unattached Disks' sheet.")
        row += 1

    # 5. SQL Server licenses (deep view across VM + DB + MI)
    sql_savings_period = sum(r["PotentialSavingsPeriod"] or 0 for r in sql_rows)
    sql_savings_annual = sql_savings_period * annualize
    if sql_rows:
        row += 1
        rec.cell(row=row, column=1, value="5. SQL Server licenses").font = Font(bold=True, size=12)
        row += 1
        not_applied_sql = sum(1 for r in sql_rows
                              if r["AhbState"] in ("Not applied", "Partially applied"))
        partial_sql = sum(1 for r in sql_rows if r["AhbState"] == "Partially applied")
        dtu_sql     = sum(1 for r in sql_rows if r["AhbState"] == "Not eligible (DTU tier)")
        applied_sql = sum(1 for r in sql_rows if r["AhbState"] == "Applied")
        rec.cell(row=row, column=1,
                 value=f"  • {len(sql_rows)} SQL resource(s) total "
                       f"(SQL on VM + Azure SQL DB + Managed Instance).")
        row += 1
        rec.cell(row=row, column=1,
                 value=f"  • {applied_sql} already AHB-applied; {not_applied_sql} "
                       f"with paid SQL licence rows.")
        row += 1
        if partial_sql:
            rec.cell(row=row, column=1,
                     value=f"  • {partial_sql} VM(s) PARTIALLY applied (mixed AHB / "
                           f"paid licence rows in {period_short}) — likely toggled "
                           f"mid-period. Verify AHB is set on every replica/region.")
            row += 1
        if dtu_sql:
            rec.cell(row=row, column=1,
                     value=f"  • {dtu_sql} Azure SQL DB(s) on DTU tier — AHB "
                           f"is INELIGIBLE on DTU. Migrate to vCore tier first, "
                           f"then enable AHB.")
            row += 1
        if sql_savings_period > 0:
            rec.cell(row=row, column=1,
                     value=f"  • Saveable SQL licence cost in {period_short}: "
                           f"{sql_savings_period:,.2f} {currency}.")
            row += 1
            rec.cell(row=row, column=1,
                     value=f"  • Annualized savings if AHB applied across all "
                           f"eligible SQL workloads: "
                           f"{sql_savings_annual:,.2f} {currency}/year.")
            rec.cell(row=row, column=1).font = accent
            row += 1
        rec.cell(row=row, column=1,
                 value="  • Action: enable AHB on SQL Server VMs (Std/Ent with "
                       "active Software Assurance) via 'License type' on the VM. "
                       "For Azure SQL DB/MI on vCore, set 'Azure Hybrid Benefit' "
                       "to Yes (on the database/instance Compute + storage blade). "
                       "See 'SQL Licenses' and 'SQL Opportunities' sheets.")
        row += 1

    rec.column_dimensions["A"].width = 110

    # =========================================================
    # Sheet: RI Candidates
    # =========================================================
    if ri_top:
        ric = wb.create_sheet("RI Candidates")
        ric["A1"] = (f"RI candidates — {len(ri_top)} (VmSize × region) groups "
                     f"on-demand in {period_label}")
        ric["A1"].font = big_bold
        ric.merge_cells("A1:H1")
        if observed_ri_discount is not None:
            ric["A2"] = (f"Observed customer RI discount: "
                         f"{observed_ri_discount * 100:.1f}% off list.")
            ric["A2"].font = Font(italic=True)
        ri_headers = ["VmSize", "Location", "VmCount", "TotalHours",
                      "AvgHoursPerVm", "Coverage",
                      "ActualCost", "ListCost",
                      "ProjectedSavingsPeriod", "ProjectedSavingsAnnual",
                      "Currency"]
        set_header(ric, 4, ri_headers)
        ri_money = {"ActualCost", "ListCost",
                    "ProjectedSavingsPeriod", "ProjectedSavingsAnnual"}
        for i, r in enumerate(ri_top, start=5):
            for c_idx, h in enumerate(ri_headers, start=1):
                cell = ric.cell(row=i, column=c_idx, value=r.get(h))
                if h in ri_money and r.get(h) is not None:
                    cell.number_format = money_fmt
                elif h in {"TotalHours", "AvgHoursPerVm"} and r.get(h) is not None:
                    cell.number_format = "#,##0.0"
        autosize(ric, ri_headers, ri_top)
        ric.freeze_panes = "A5"

    # =========================================================
    # Sheet: Spot Candidates (bursty/dev VMs eligible for Spot)
    # =========================================================
    if spot_candidates:
        spc = wb.create_sheet("Spot Candidates")
        spc_total = sum(c.get("ProjectedSavingsPeriod") or 0 for c in spot_candidates)
        spc["A1"] = (f"Spot candidates — {len(spot_candidates)} (VmSize × region) "
                     f"groups of bursty/dev VMs in {period_label} "
                     f"(potential savings: {spc_total:,.2f} {currency})")
        spc["A1"].font = big_bold
        spc.merge_cells("A1:H1")
        spc["A2"] = (f"Reference rate: {int(vms_mod.SPOT_REFERENCE_DISCOUNT*100)}% off list. "
                     f"Spot VMs can be evicted with 30-second notice — only move "
                     f"fault-tolerant workloads (batch, dev/test, stateless services).")
        spc["A2"].font = Font(italic=True)
        spc.merge_cells("A2:H2")
        spot_headers = ["VmSize", "Location", "VmCount", "TotalHours",
                        "AvgHoursPerVm", "Coverage", "ActualCost",
                        "ProjectedSavingsPeriod", "Currency"]
        set_header(spc, 4, spot_headers)
        spot_money = {"ActualCost", "ProjectedSavingsPeriod"}
        for i, r in enumerate(spot_candidates, start=5):
            for c_idx, h in enumerate(spot_headers, start=1):
                cell = spc.cell(row=i, column=c_idx, value=r.get(h))
                if h in spot_money and r.get(h) is not None:
                    cell.number_format = money_fmt
                elif h in {"TotalHours", "AvgHoursPerVm"} and r.get(h) is not None:
                    cell.number_format = "#,##0.0"
        autosize(spc, spot_headers, spot_candidates)
        spc.freeze_panes = "A5"

    # =========================================================
    # Sheet: AHB Opportunities (top by savings)
    # =========================================================
    if ahb_top:
        ato = wb.create_sheet("AHB Opportunities")
        ato_total = sum(r["PotentialSavingsPeriod"] or 0 for r in ahb_top)
        ato["A1"] = (f"AHB opportunities — {len(ahb_top)} resource(s) where AHB "
                     f"is NOT applied. Period: {period_label}, "
                     f"savings = {ato_total:,.2f} {currency} "
                     f"(~{ato_total * annualize:,.2f} {currency}/year).")
        ato["A1"].font = big_bold
        ato.merge_cells("A1:H1")
        ato_headers = ["Workload", "AhbState", "Edition", "Name",
                       "ResourceGroup", "SubscriptionId", "Location",
                       "Currency", "ComputeCostInPeriod", "LicenseCostInPeriod",
                       "TotalCostInPeriod", "PotentialSavingsPeriod",
                       "ResourceId"]
        set_header(ato, 3, ato_headers)
        ato_money = {"ComputeCostInPeriod", "LicenseCostInPeriod",
                     "TotalCostInPeriod", "PotentialSavingsPeriod"}
        for i, r in enumerate(ahb_top, start=4):
            for c_idx, h in enumerate(ato_headers, start=1):
                cell = ato.cell(row=i, column=c_idx, value=r.get(h))
                if h in ato_money and r.get(h) is not None:
                    cell.number_format = money_fmt
        autosize(ato, ato_headers, ahb_top)
        ato.freeze_panes = "A4"
        ato.auto_filter.ref = f"A3:{get_column_letter(len(ato_headers))}{3 + len(ahb_top)}"

    # =========================================================
    # Sheet: SQL Opportunities (top by saveable license cost)
    # =========================================================
    if sql_top:
        sto = wb.create_sheet("SQL Opportunities")
        sto_total = sum(r["PotentialSavingsPeriod"] or 0 for r in sql_top)
        sto["A1"] = (f"SQL license opportunities — {len(sql_top)} resource(s) "
                     f"with paid SQL licence rows. Period: {period_label}, "
                     f"savings = {sto_total:,.2f} {currency} "
                     f"(~{sto_total * annualize:,.2f} {currency}/year).")
        sto["A1"].font = big_bold
        sto.merge_cells("A1:H1")
        sto_headers = ["Workload", "Tier", "AhbState", "Name",
                       "ResourceGroup", "SubscriptionId", "Location",
                       "Currency", "ComputeCostInPeriod", "StorageCostInPeriod",
                       "LicenseCostInPeriod", "OtherCostInPeriod",
                       "TotalCostInPeriod", "PotentialSavingsPeriod",
                       "ResourceId", "Notes"]
        set_header(sto, 3, sto_headers)
        sto_money = {"ComputeCostInPeriod", "StorageCostInPeriod",
                     "LicenseCostInPeriod", "OtherCostInPeriod",
                     "TotalCostInPeriod", "PotentialSavingsPeriod"}
        for i, r in enumerate(sql_top, start=4):
            for c_idx, h in enumerate(sto_headers, start=1):
                cell = sto.cell(row=i, column=c_idx, value=r.get(h))
                if h in sto_money and r.get(h) is not None:
                    cell.number_format = money_fmt
        autosize(sto, sto_headers, sql_top)
        sto.freeze_panes = "A4"
        sto.auto_filter.ref = f"A3:{get_column_letter(len(sto_headers))}{3 + len(sql_top)}"

    # =========================================================
    # Dashboard sheet (assembled last so chart refs already exist)
    # =========================================================
    dash = wb.create_sheet("Dashboard", 0)  # insert as first sheet

    dash["A1"] = "FinOps Dashboard"
    dash["A1"].font = huge_bold
    dash.merge_cells("A1:H1")

    dash["A2"] = f"Billing period: {period_label}"
    dash["A2"].font = big_bold
    dash.merge_cells("A2:H2")

    dash["A3"] = (f"Time grain: 'period' = {period_label}. "
                  f"'Monthly' = full-month run rate (rate × {HOURS_PER_MONTH}h). "
                  f"'Annual' = period × 365/days extrapolation.")
    dash["A3"].font = Font(italic=True, color="555555")
    dash.merge_cells("A3:H3")

    # ---- Headline numbers
    total_vm_period_cost   = sum(r["ActualCostInPeriod"] or 0 for r in vm_rows)
    total_vm_period_save   = sum(r["ActualSavings"]      or 0 for r in vm_rows)
    total_disk_monthly     = sum(r["CustomerMonthlyCost"] or 0 for r in disk_rows)
    total_ahb_period_save  = sum(r["PotentialSavingsPeriod"] or 0 for r in ahb_rows)
    total_ri_period_save   = sum(c.get("ProjectedSavingsPeriod") or 0 for c in ri_top)
    # Combined opportunity (rough — overlap is small because RI/AHB scopes
    # are different).
    total_opp_period       = total_ahb_period_save + total_ri_period_save \
                             + (total_disk_monthly * (period.get('days') or 30) / HOURS_PER_MONTH * HOURS_PER_MONTH) \
                             if False else (total_ahb_period_save + total_ri_period_save)
    # Note: 'unattached disks (monthly run-rate)' is already a monthly number,
    # not period-scaled. Adding it to a period total would double-count units.
    # Keep period total as AHB + RI only; show disks separately.

    # Pre-compute commitment-cost roll-ups and estimated savings.
    realized_neg = realized_savings.get("Negotiated discount (no commitment)", 0.0)
    covered_ri   = cost_covered.get("Reservation",   0.0)
    covered_sp   = cost_covered.get("Savings Plan",  0.0)
    covered_spot = cost_covered.get("Spot",          0.0)
    covered_total = covered_ri + covered_sp + covered_spot

    # Estimate RI/SP savings using the customer's OWN observed RI discount.
    if observed_ri_discount and 0 < observed_ri_discount < 1:
        ri_savings_factor = observed_ri_discount / (1 - observed_ri_discount)
    else:
        ri_savings_factor = 0.0
    estimated_ri_savings = covered_ri * ri_savings_factor
    estimated_sp_savings = covered_sp * ri_savings_factor

    # AHB realized: how much AHB is currently APPLIED on those rows that CAN
    # have AHB. The AHB scan tracks per-resource license cost; we approximate
    # 'realized AHB' as: license cost the customer would pay if they removed
    # AHB from currently-AHB-applied resources. That's hard to compute
    # without a counterfactual, so we report 'AHB applied to N resources'
    # informationally and quantify the OPPORTUNITY (not-applied) precisely.
    ahb_resources_applied = sum(1 for r in ahb_rows
                                if (r.get("AhbState") or "").startswith("Applied"))
    ahb_resources_review  = sum(1 for r in ahb_rows
                                if (r.get("AhbState") or "").startswith("Review"))

    realized_total = realized_neg + estimated_ri_savings + estimated_sp_savings + covered_spot
    additional_total = total_ahb_period_save + total_ri_period_save

    # ----- SQL roll-ups
    # Total SQL savings (across all SQL workloads) — the SQL scan is
    # authoritative.
    total_sql_period_save = sum(r["PotentialSavingsPeriod"] or 0 for r in sql_rows)
    # AHB scan already counts SQL Server on VM in `total_ahb_period_save`,
    # so for the combined headline we only ADD the Azure SQL DB + MI portion.
    sql_db_mi_period_save = sum(r["PotentialSavingsPeriod"] or 0
                                for r in sql_rows
                                if r["Workload"] != "SQL Server on VM")
    total_sql_period_cost = sum(r["TotalCostInPeriod"] or 0 for r in sql_rows)
    total_sql_license_period = sum(r["LicenseCostInPeriod"] or 0 for r in sql_rows)
    sql_resources_partial = sum(1 for r in sql_rows
                                if r["AhbState"] == "Partially applied")
    sql_resources_dtu = sum(1 for r in sql_rows
                            if r["AhbState"] == "Not eligible (DTU tier)")

    cur = currency
    pct_of_spend = lambda v: (v / total_period_cost * 100) if total_period_cost else 0.0

    # ----- Layout helpers
    cur_row = 5
    section_fg = "FFFFFF"

    def write_section_header(title: str):
        nonlocal cur_row
        c = dash.cell(row=cur_row, column=1, value=title)
        c.font = Font(bold=True, size=13, color=section_fg)
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        # Span across all data columns so the bar reads cleanly.
        # Cols A..D = label / current / would-pay / savings.
        # Also fill cols B..D with the section colour so the bar is
        # uninterrupted (merge_cells alone doesn't paint the right cells in
        # all viewers).
        for col in range(2, 5):
            dash.cell(row=cur_row, column=col).fill = section_fill
        dash.merge_cells(start_row=cur_row, start_column=1,
                         end_row=cur_row, end_column=4)
        cur_row += 1

    def write_metric(label: str, value: float, unit: str, *, kind: str = "neutral"):
        """Single-value metric. Layout: A=label, B=value, C=unit.

        kind ∈ {'realized','opportunity','total','context','neutral'}
        """
        nonlocal cur_row
        dash.cell(row=cur_row, column=1, value=label)
        c = dash.cell(row=cur_row, column=2, value=round(value, 2))
        c.number_format = money_fmt if unit != "% of spend" else '0.0"%"'
        dash.cell(row=cur_row, column=3, value=unit)
        if kind == "total":
            dash.cell(row=cur_row, column=1).font = Font(bold=True, size=12, color="0E5C2F")
            c.font = Font(bold=True, size=12, color="0E5C2F")
        elif kind == "realized":
            dash.cell(row=cur_row, column=1).font = Font(bold=True, color="217346")
            c.font = Font(bold=True, color="217346")
        elif kind == "opportunity":
            # Opportunity is the headline — bigger, bolder, taller row.
            dash.cell(row=cur_row, column=1).font = Font(bold=True, size=13, color="0070C0")
            c.font = Font(bold=True, size=13, color="0070C0")
            dash.row_dimensions[cur_row].height = 22
        elif kind == "context":
            dash.cell(row=cur_row, column=1).font = Font(color="555555")
            c.font = Font(color="555555")
        cur_row += 1

    def write_opp_columns_header():
        """Column legend for opportunity rows. Cols B/C/D = current / would-pay / savings."""
        nonlocal cur_row
        for c_idx, label in enumerate(
                [None, f"Currently paying ({currency})",
                 f"Would pay ({currency})", f"Savings ({currency})"],
                start=1):
            if not label:
                continue
            cell = dash.cell(row=cur_row, column=c_idx, value=label)
            cell.font = Font(bold=True, size=10, color="555555")
            cell.alignment = Alignment(horizontal="right")
        cur_row += 1

    def write_opportunity_metric(label: str, current: float, would_pay: float,
                                 savings: float, *, kind: str = "opportunity"):
        """3-number opportunity row. A=label, B=current, C=would-pay, D=savings."""
        nonlocal cur_row
        dash.cell(row=cur_row, column=1, value=label)
        b = dash.cell(row=cur_row, column=2, value=round(current, 2))
        c = dash.cell(row=cur_row, column=3, value=round(would_pay, 2))
        d = dash.cell(row=cur_row, column=4, value=round(savings, 2))
        for cell in (b, c, d):
            cell.number_format = money_fmt
        if kind == "total":
            font = Font(bold=True, size=12, color="0E5C2F")
            for cell in (dash.cell(row=cur_row, column=1), b, c, d):
                cell.font = font
        elif kind == "opportunity":
            label_font = Font(bold=True, size=13, color="0070C0")
            value_font = Font(bold=True, size=13, color="0070C0")
            dash.cell(row=cur_row, column=1).font = label_font
            for cell in (b, c, d):
                cell.font = value_font
            dash.row_dimensions[cur_row].height = 22
        cur_row += 1

    def write_note(text: str):
        nonlocal cur_row
        c = dash.cell(row=cur_row, column=1, value=text)
        c.font = Font(italic=True, color="555555", size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        dash.merge_cells(start_row=cur_row, start_column=1,
                         end_row=cur_row, end_column=4)
        cur_row += 1

    def blank_row():
        nonlocal cur_row
        cur_row += 1

    # =================== Total spend ===================
    write_section_header("TOTAL SPEND")
    write_metric(f"All services, all-in ({period_short})", total_period_cost,         cur,             kind="total")
    write_metric(f"At LIST (PAYG) ({period_short})",       total_period_list,         cur,             kind="context")
    blank_row()

    # ----- Pre-compute SP opportunity using a public reference rate.
    # SP and RI cover the same on-demand VMs — buy ONE OR THE OTHER, not both.
    # 1-yr Compute Savings Plan ≈ 17% off list (public reference).
    SP_REFERENCE_DISCOUNT = 0.17
    if observed_ri_discount and 0 < observed_ri_discount < 1:
        implied_on_demand_list = total_ri_period_save / observed_ri_discount
    else:
        implied_on_demand_list = 0.0
    sp_opportunity_period = implied_on_demand_list * SP_REFERENCE_DISCOUNT

    # =================== Roll-up (front and center) ===================
    # The 'best' opportunity is RI (typically the deeper discount), so
    # combined opportunity = AHB + RI/SP + Spot + waste. Don't double-count
    # SP against RI (they target the same workloads — pick one). Spot covers
    # a different pool (bursty/dev VMs that aren't RI candidates), so it's
    # additive.
    spot_opportunity_for_headline = sum(c.get("ProjectedSavingsPeriod") or 0
                                        for c in spot_candidates)
    waste_disks_period = total_disk_monthly  # already monthly; treat as period
    combined_opportunity_period = (total_ahb_period_save
                                   + total_ri_period_save
                                   + spot_opportunity_for_headline
                                   + sql_db_mi_period_save
                                   + waste_disks_period)

    # ----- 'Currently paying' baselines for each opportunity.
    # For each action we report:
    #   current   = what the customer pays today on the affected meters
    #   would_pay = what they'd pay after taking the action
    #   savings   = current - would_pay (matches the existing OPPORTUNITY number)
    #
    # Two action shapes:
    #   Discount actions (RI/SP, Spot) — current = full VM cost on the
    #     candidate set; would_pay = current minus the discount.
    #   Elimination actions (AHB, deleting disks) — current = the addressable
    #     meter cost only (the licence surcharge or the unattached-disk
    #     storage charge); would_pay = $0; savings = current.
    ri_current   = sum(c.get("ActualCost") or 0 for c in ri_top)
    ri_would     = max(ri_current - total_ri_period_save, 0.0)

    spot_current = sum(c.get("ActualCost") or 0 for c in spot_candidates)
    spot_would   = max(spot_current - spot_opportunity_for_headline, 0.0)

    # AHB on Windows + SQL Server on VM: the licence surcharge meter is what
    # disappears when AHB is enabled. Compute / OS keeps being billed on
    # separate meters that AHB doesn't touch.
    ahb_current  = total_ahb_period_save
    ahb_would    = 0.0

    # SQL AHB on Azure SQL DB / MI: the '... SQL License' meter disappears.
    # Compute and storage rows on the same DB stay (and are not in this
    # number).
    sql_dbmi_current = sql_db_mi_period_save
    sql_dbmi_would   = 0.0

    disk_current = waste_disks_period
    disk_would   = 0.0

    combined_current = total_period_cost
    combined_would   = max(total_period_cost - combined_opportunity_period, 0.0)

    write_section_header("⭐ HEADLINE — WHAT YOU SHOULD ACT ON")
    write_opp_columns_header()
    write_opportunity_metric(
        f"COMBINED OPPORTUNITY in {period_short} (RI + Spot + AHB + SQL + waste)",
        combined_current, combined_would, combined_opportunity_period,
        kind="opportunity")
    write_opportunity_metric(
        f"  └ via Reservations (or Savings Plans) — always-on VMs",
        ri_current, ri_would, total_ri_period_save)
    if spot_opportunity_for_headline > 0:
        write_opportunity_metric(
            f"  └ via Spot — bursty/dev VMs",
            spot_current, spot_would, spot_opportunity_for_headline)
    write_opportunity_metric(
        f"  └ via Azure Hybrid Benefit (Windows + SQL on VM)",
        ahb_current, ahb_would, total_ahb_period_save)
    if sql_db_mi_period_save > 0:
        write_opportunity_metric(
            f"  └ via SQL AHB on Azure SQL DB / Managed Instance",
            sql_dbmi_current, sql_dbmi_would, sql_db_mi_period_save)
    if waste_disks_period > 0:
        write_opportunity_metric(
            f"  └ via deleting unattached disks",
            disk_current, disk_would, waste_disks_period)
    write_metric(f"REALIZED savings already captured in {period_short}",
                                                                      realized_total,              cur,             kind="realized")
    blank_row()

    # =================== Reservations ===================
    write_section_header("RESERVATIONS (RI)")
    write_metric(f"OPPORTUNITY in {period_short} (additional, on-demand VMs)",
                                                                      total_ri_period_save,        cur,             kind="opportunity")
    write_metric(f"REALIZED savings in {period_short} (estimated*)", estimated_ri_savings,        cur,             kind="realized")
    write_metric(f"Cost going through RIs in {period_short}",         covered_ri,                  cur,             kind="context")
    write_metric(f"% of total spend covered by RIs",                  pct_of_spend(covered_ri),    "% of spend",    kind="context")
    blank_row()

    # =================== Savings Plans ===================
    write_section_header("SAVINGS PLANS (SP)")
    write_metric(f"OPPORTUNITY in {period_short} (same on-demand VMs, SP reference rate {int(SP_REFERENCE_DISCOUNT*100)}%)",
                                                                      sp_opportunity_period,       cur,             kind="opportunity")
    write_metric(f"REALIZED savings in {period_short} (estimated*)", estimated_sp_savings,        cur,             kind="realized")
    write_metric(f"Cost going through SPs in {period_short}",         covered_sp,                  cur,             kind="context")
    write_metric(f"% of total spend covered by SPs",                  pct_of_spend(covered_sp),    "% of spend",    kind="context")
    write_note("Note: Buying SP would replace the RI opportunity above — these are alternative ways to "
               "cover the same on-demand VMs. Pick the one that fits your workload pattern.")
    blank_row()

    # =================== Azure Hybrid Benefit (AHB) ===================
    write_section_header("AZURE HYBRID BENEFIT (AHB)")
    write_metric(f"OPPORTUNITY in {period_short} (additional)",       total_ahb_period_save,       cur,             kind="opportunity")
    write_metric(f"REALIZED — # of resources with AHB applied",       ahb_resources_applied,       "resources",     kind="realized")
    if ahb_resources_review:
        write_metric(f"For review (Azure SQL — vCore vs license-included)",
                                                                      ahb_resources_review,         "resources",     kind="context")
    blank_row()

    # =================== SQL Server licenses (deep view) ===================
    write_section_header("SQL SERVER LICENSES")
    write_metric(f"Total SQL spend in {period_short} (compute + storage + license)",
                                                                      total_sql_period_cost,        cur,             kind="context")
    write_metric(f"Of which is SQL licence cost",                     total_sql_license_period,    cur,             kind="context")
    write_metric(f"OPPORTUNITY in {period_short} (full SQL — VM + DB + MI)",
                                                                      total_sql_period_save,        cur,             kind="opportunity")
    if sql_db_mi_period_save > 0:
        write_metric(f"  └ Azure SQL DB / MI (additional vs AHB section)",
                                                                      sql_db_mi_period_save,        cur,             kind="opportunity")
    if sql_resources_partial:
        write_metric(f"# VMs with PARTIALLY-applied SQL AHB (toggle gap)",
                                                                      sql_resources_partial,        "VMs",           kind="context")
    if sql_resources_dtu:
        write_metric(f"# Azure SQL DBs on DTU tier (AHB ineligible — migrate to vCore)",
                                                                      sql_resources_dtu,            "DBs",           kind="context")
    write_note("Full SQL opportunity above includes SQL Server on VM (also counted in the AHB section). "
               "The combined headline only adds the Azure SQL DB / MI portion to avoid double-counting. "
               "See 'SQL Licenses' and 'SQL Opportunities' sheets for the per-resource breakdown.")
    blank_row()

    # =================== MCA Negotiated Discount ===================
    write_section_header("MCA NEGOTIATED DISCOUNT")
    write_metric(f"REALIZED savings in {period_short}",               realized_neg,                cur,             kind="realized")
    write_metric(f"% of total spend (effective discount)",            pct_of_spend(realized_neg),  "% of spend",    kind="context")
    blank_row()

    # =================== Spot ===================
    spot_opportunity_period = sum(c.get("ProjectedSavingsPeriod") or 0
                                  for c in spot_candidates)
    write_section_header("SPOT")
    if spot_opportunity_period > 0:
        write_metric(f"OPPORTUNITY in {period_short} (bursty/dev VMs at "
                     f"{int(vms_mod.SPOT_REFERENCE_DISCOUNT*100)}% reference rate)",
                                                                      spot_opportunity_period,     cur,             kind="opportunity")
        write_metric(f"  └ # bursty/dev VM groups eligible",          len(spot_candidates),         "groups",        kind="context")
    write_metric(f"Cost going through Spot in {period_short}",        covered_spot,                cur,             kind="context")
    write_metric(f"% of total spend on Spot",                         pct_of_spend(covered_spot),  "% of spend",    kind="context")
    if spot_opportunity_period > 0:
        write_note("Spot opportunity = bursty/dev VMs (avg <50% of period hours) re-priced at the public "
                   "Spot reference rate. Move ONLY fault-tolerant workloads — VMs can be evicted with "
                   "30-second notice. See 'Spot Candidates' sheet for the per-group breakdown.")
    blank_row()

    # =================== ADDITIONAL OPPORTUNITIES (waste) ===================
    # Resources with measurable waste — i.e., things you're paying for that
    # provide no value. Today: unattached disks. Future: idle VMs, orphaned
    # NICs/PIPs, oversized disks, etc.
    write_section_header("ADDITIONAL OPPORTUNITIES (WASTE)")
    if disk_rows:
        write_metric(f"OPPORTUNITY — unattached disks ({len(disk_rows)} disk(s), savings if all deleted)",
                                                                      total_disk_monthly,          f"{cur}/month",   kind="opportunity")
        write_metric(f"  See 'Unattached Disks' sheet for the full list",
                                                                      len(disk_rows),               "disks",         kind="context")
    else:
        write_note("No unattached disks found in scope. Add other waste signals here as needed "
                   "(idle VMs, orphaned NICs / public IPs, oversized disks, stopped-deallocated VMs "
                   "still incurring storage cost, etc.).")
    blank_row()
    write_note(f"* RI / SP realized savings are ESTIMATED using your observed RI discount of "
               f"{(observed_ri_discount * 100 if observed_ri_discount else 0):.1f}% (derived from existing "
               "RI-covered VMs). FOCUS records commitment-covered usage with ListCost=0, so exact per-row "
               "(list − effective) arithmetic doesn't apply on those rows. SP opportunity uses a public "
               f"reference rate of {int(SP_REFERENCE_DISCOUNT*100)}% (1-yr Compute SP). MCA negotiated and "
               "AHB opportunity numbers are exact (per-row arithmetic on the billing CSV).")

    dash.column_dimensions["A"].width = 64
    dash.column_dimensions["B"].width = 20
    dash.column_dimensions["C"].width = 18
    dash.column_dimensions["D"].width = 18

    # ---- Charts
    # Place charts to the right of the section table (col E onward)
    chart_anchor_row = 5

    # Chart 0: Pie — TOTAL spend by service (the headline view)
    if sv_end >= sv_start:
        pie0 = PieChart()
        pie0.title = f"Total spend by service ({period_short})"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=sv_start, max_row=sv_end)
        data   = Reference(cd, min_col=2, max_col=2,
                           min_row=sv_start - 1, max_row=sv_end)
        pie0.add_data(data, titles_from_data=True)
        pie0.set_categories(labels)
        pie0.height = 9
        pie0.width  = 18
        pie0.dataLabels = DataLabelList(
            showPercent=True, showCatName=False, showVal=False,
            showSerName=False, showLegendKey=False, showBubbleSize=False)
        dash.add_chart(pie0, f"E{chart_anchor_row}")

    # Chart 1: Pie — VM cost by benefit category
    if cat_end >= cat_start:
        pie1 = PieChart()
        pie1.title = f"VM cost by benefit category ({period_short})"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=cat_start, max_row=cat_end)
        data   = Reference(cd, min_col=3, max_col=3,
                           min_row=cat_start - 1, max_row=cat_end)  # include header
        pie1.add_data(data, titles_from_data=True)
        pie1.set_categories(labels)
        pie1.height = 8
        pie1.width  = 16
        pie1.dataLabels = DataLabelList(
            showPercent=True, showCatName=False, showVal=False,
            showSerName=False, showLegendKey=False, showBubbleSize=False)
        dash.add_chart(pie1, f"E{chart_anchor_row + 20}")

    # Chart 2: Pie — AHB by license cost (excludes the no-action bucket).
    # Cost-weighting means rows with $0 license cost (Linux, AHB-applied,
    # Azure SQL on vCore, etc.) are absent, so the pie only shows the
    # actionable spend split.
    if ahb_pie_first is not None and ahb_pie_last is not None:
        pie2 = PieChart()
        pie2.title = (f"Where AHB-eligible license $ goes ({period_short}, "
                      f"actionable only)")
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=ahb_pie_first, max_row=ahb_pie_last)
        # Column 3 is now License cost (we removed the State column).
        data   = Reference(cd, min_col=3, max_col=3,
                           min_row=ahb_pie_first - 1, max_row=ahb_pie_last)
        pie2.add_data(data, titles_from_data=True)
        pie2.set_categories(labels)
        pie2.height = 8
        pie2.width  = 16
        pie2.dataLabels = DataLabelList(
            showPercent=True, showCatName=False, showVal=False,
            showSerName=False, showLegendKey=False, showBubbleSize=False)
        dash.add_chart(pie2, f"E{chart_anchor_row + 38}")

    # Chart 3: Bar — Top RI candidates (annualized savings)
    if ri_end >= ri_start:
        bar1 = BarChart()
        bar1.type  = "bar"
        bar1.style = 11
        bar1.title = f"Top RI candidates — annualized savings ({currency}/year)"
        bar1.y_axis.title = "VmSize × Region"
        bar1.x_axis.title = f"Projected annual savings ({currency}/year)"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=ri_start, max_row=ri_end)
        data   = Reference(cd, min_col=5, max_col=5,
                           min_row=ri_start - 1, max_row=ri_end)
        bar1.add_data(data, titles_from_data=True)
        bar1.set_categories(labels)
        bar1.height = 12
        bar1.width  = 22
        dash.add_chart(bar1, f"A{chart_anchor_row + 56}")

    # Chart 4: Bar — Top AHB opportunities (annualized savings)
    if ahb_top_end >= ahb_top_start:
        bar2 = BarChart()
        bar2.type  = "bar"
        bar2.style = 12
        bar2.title = f"Top AHB opportunities — annualized savings ({currency}/year)"
        bar2.y_axis.title = "Resource"
        bar2.x_axis.title = f"Annualized savings ({currency}/year)"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=ahb_top_start, max_row=ahb_top_end)
        data   = Reference(cd, min_col=5, max_col=5,
                           min_row=ahb_top_start - 1, max_row=ahb_top_end)
        bar2.add_data(data, titles_from_data=True)
        bar2.set_categories(labels)
        bar2.height = 12
        bar2.width  = 22
        dash.add_chart(bar2, f"A{chart_anchor_row + 80}")

    # Chart 5: Bar — Top unattached disks (monthly cost)
    if dk_end >= dk_start:
        bar3 = BarChart()
        bar3.type  = "bar"
        bar3.style = 13
        bar3.title = f"Top unattached disks — monthly cost ({currency}/month)"
        bar3.y_axis.title = "Disk"
        bar3.x_axis.title = f"Monthly cost ({currency}/month)"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=dk_start, max_row=dk_end)
        data   = Reference(cd, min_col=3, max_col=3,
                           min_row=dk_start - 1, max_row=dk_end)
        bar3.add_data(data, titles_from_data=True)
        bar3.set_categories(labels)
        bar3.height = 12
        bar3.width  = 22
        dash.add_chart(bar3, f"A{chart_anchor_row + 104}")

    # Chart 6: Bar — Savings realized vs additional opportunity (the
    # headline 'where we save' visual). Two clustered bars per mechanism.
    if sav_end >= sav_start:
        bar4 = BarChart()
        bar4.type  = "col"  # vertical clustered bars
        bar4.style = 26
        bar4.title = (f"Savings: REALIZED (already captured) vs ADDITIONAL "
                      f"opportunity ({period_short}, {currency})")
        bar4.y_axis.title = f"{currency} in {period_short}"
        bar4.x_axis.title = "Mechanism"
        labels = Reference(cd, min_col=1, max_col=1,
                           min_row=sav_start, max_row=sav_end)
        # Two data columns (realized, additional) with their headers, so the
        # legend reads correctly.
        data = Reference(cd, min_col=2, max_col=3,
                         min_row=sav_start - 1, max_row=sav_end)
        bar4.add_data(data, titles_from_data=True)
        bar4.set_categories(labels)
        bar4.height = 11
        bar4.width  = 22
        dash.add_chart(bar4, f"A{chart_anchor_row + 128}")

    # Save
    wb.save(out_path)


def period_days_label(period: dict) -> str:
    days = period.get("days")
    if days:
        return f"{days} days"
    return "unknown duration"


if __name__ == "__main__":
    sys.exit(main())
