#!/usr/bin/env python3
"""Scan a billing CSV for Azure Hybrid Benefit (AHB) opportunity.

Reports every resource where AHB CAN apply, whether it IS applied, and the
cost being paid for non-AHB licenses (= savings if AHB were enabled).

Workload coverage
=================
1. Windows on Azure VM
   - AHB ON  : VM has compute meter rows but NO 'Windows' surcharge meter row.
   - AHB OFF : VM has a surcharge row in subcategory '... Windows'.
2. SQL Server on Azure VM (BYOL via AHB)
   - AHB ON  : meter category 'Virtual Machines Licenses', subcategory
               'SQL Server Azure Hybrid Benefit'.
   - AHB OFF : same category, subcategory 'SQL Server Standard' or
               'SQL Server Enterprise'.
   - N/A     : 'SQL Server Developer Edition' or 'SQL Server Express Edition'
               (free) — surfaced for completeness.
3. Azure SQL Database / Managed Instance
   - The license vs vCore split is harder to detect from billing alone. We
     surface every Azure SQL resource with cost breakdown so the customer
     can review whether AHB (vCore + BYOL) is in use.
4. Red Hat / SUSE on Azure VM
   - AHB doesn't apply (Microsoft-only). Surfaced for awareness — a separate
     BYOS / license-mobility decision.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

HOURS_PER_MONTH = 730


# ---------- CSV schemas ----------

LEGACY_SCHEMA = {
    "name":             "legacy",
    "category":         "meterCategory",
    "subcategory":      "meterSubCategory",
    "meter":            "meterName",
    "region":           "resourceLocation",
    "service":          "consumedService",
    "service_family":   "serviceFamily",
    "currency":         "pricingCurrency",
    "resource_id":      "resourceId",
    "quantity":         "quantity",
    "cost":             "costInBillingCurrency",
    "list_cost":        "paygCostInBillingCurrency",
    "period_start":     "billingPeriodStartDate",
    "period_end":       "billingPeriodEndDate",
    "unit":             "unitOfMeasure",
}
FOCUS_SCHEMA = {
    "name":             "FOCUS",
    "category":         "x_SkuMeterCategory",
    "subcategory":      "x_SkuMeterSubcategory",
    "meter":            "x_SkuMeterName",
    "region":           "RegionId",
    "service":          "ServiceName",
    "service_family":   "x_SkuServiceFamily",
    "currency":         "BillingCurrency",
    "resource_id":      "ResourceId",
    "quantity":         "PricingQuantity",
    "cost":             "EffectiveCost",
    "list_cost":        "ListCost",
    "period_start":     "BillingPeriodStart",
    "period_end":       "BillingPeriodEnd",
    "unit":             "PricingUnit",
}


def _detect_schema(fieldnames: list[str]) -> Optional[dict]:
    for schema in (LEGACY_SCHEMA, FOCUS_SCHEMA):
        cols = {schema[k] for k in (
            "category", "subcategory", "meter", "service",
            "currency", "resource_id", "quantity", "cost",
        )}
        if cols.issubset(fieldnames):
            return schema
    return None


def _to_float(v: Optional[str]) -> Optional[float]:
    if not v:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _parse_resource_id(rid: str) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """Returns (subId, rg, type, name) from an ARM resource id."""
    if not rid:
        return (None, None, None, None)
    parts = rid.split("/")
    sub = rg = name = rtype = None
    try:
        if len(parts) > 2 and parts[1].lower() == "subscriptions":
            sub = parts[2]
        if len(parts) > 4 and parts[3].lower() == "resourcegroups":
            rg = parts[4]
        if len(parts) > 7:
            rtype = f"{parts[6]}/{parts[7]}"  # provider/type
        name = parts[-1] if parts[-1] else None
    except IndexError:
        pass
    return (sub, rg, rtype, name)


# ---------- Workload classification ----------

@dataclass
class AhbAggregation:
    resource_id: str
    name: str
    resource_group: str
    subscription_id: str
    resource_type: str = ""
    location: str = ""
    currency: str = ""
    workload: str = ""              # 'Windows on VM' / 'SQL Server on VM' / ...
    ahb_state: str = ""             # 'Applied' / 'Not applied' / 'N/A' / 'Review'
    edition: str = ""               # 'Standard' / 'Enterprise' / 'Developer' / 'Express' / ''
    # Per-bucket cost: helps explain the determination.
    compute_cost: float = 0.0       # base VM/SQL compute (always charged)
    license_cost: float = 0.0       # license surcharge (= savings if AHB applied)
    total_cost: float = 0.0
    quantity_total: float = 0.0
    notes: list[str] = field(default_factory=list)


def aggregate_ahb(path: Path) -> tuple[dict[str, AhbAggregation], dict]:
    size_mb = path.stat().st_size / (1024 * 1024)
    print(f"Loading AHB-eligible billing from '{path}' ({size_mb:,.1f} MB)...")
    t0 = time.perf_counter()

    # Pre-filter: keep lines mentioning Windows VM series, SQL, RHEL, SUSE.
    KEEP_TOKENS = (
        "Virtual Machines",      # any VM compute or license row
        "SQL Database",
        "SQL Managed Instance",
        "SQL Server",
        "Red Hat",
        "SUSE",
    )
    candidate_lines: list[str] = []
    rows_scanned = 0
    progress_every = 250_000

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        header_line = f.readline()
        if not header_line:
            raise RuntimeError("Usage CSV is empty.")
        candidate_lines.append(header_line)
        for line in f:
            rows_scanned += 1
            if any(t in line for t in KEEP_TOKENS):
                candidate_lines.append(line)
            if rows_scanned % progress_every == 0:
                elapsed = time.perf_counter() - t0
                print(f"  scanned {rows_scanned:,} rows, "
                      f"{len(candidate_lines) - 1:,} candidates so far "
                      f"({elapsed:.1f}s)...")

    elapsed = time.perf_counter() - t0
    print(f"  pre-filter complete: {rows_scanned:,} rows scanned, "
          f"{len(candidate_lines) - 1:,} candidates kept in {elapsed:.1f}s.")

    reader = csv.DictReader(candidate_lines)
    schema = _detect_schema(list(reader.fieldnames or []))
    if schema is None:
        raise RuntimeError(
            "Usage CSV is neither a legacy 'Detail_BillingProfile_*' nor a "
            "FOCUS export. Required columns not found."
        )
    print(f"  detected schema: {schema['name']}")

    cat_col   = schema["category"]
    sub_col   = schema["subcategory"]
    rid_col   = schema["resource_id"]
    qty_col   = schema["quantity"]
    cost_col  = schema["cost"]
    cur_col   = schema["currency"]
    reg_col   = schema["region"]
    ps_col    = schema.get("period_start")
    pe_col    = schema.get("period_end")
    fields = set(reader.fieldnames or [])
    if ps_col not in fields: ps_col = None
    if pe_col not in fields: pe_col = None

    # Track per-VM whether we've seen a Windows surcharge row, so a VM that
    # has only the (clean) compute meter is treated as "AHB applied".
    vm_has_surcharge: dict[str, bool] = {}
    vm_has_compute:   dict[str, bool] = {}

    aggregations: dict[str, AhbAggregation] = {}
    period_start_str: Optional[str] = None
    period_end_str:   Optional[str] = None

    for row in reader:
        cat = row.get(cat_col) or ""
        sub = row.get(sub_col) or ""
        rid = row.get(rid_col) or ""
        if not rid:
            continue

        # Capture period from any matching row.
        if period_start_str is None and ps_col:
            v = (row.get(ps_col) or "").strip()
            if v:
                period_start_str = v
        if period_end_str is None and pe_col:
            v = (row.get(pe_col) or "").strip()
            if v:
                period_end_str = v

        # Classify the row.
        workload = ""
        ahb_state = ""
        edition = ""
        is_license_row = False
        is_compute_row = False

        # ---- Windows VM (surcharge meter or clean compute) ----
        if cat == "Virtual Machines":
            if "/virtualMachines/" in rid or "/virtualmachines/" in rid.lower():
                if sub.endswith(" Windows"):
                    workload = "Windows on VM"
                    ahb_state = "Not applied"
                    is_license_row = True
                else:
                    # Clean compute row — track for AHB-applied inference.
                    rid_l = rid.lower()
                    vm_has_compute[rid_l] = True
                    is_compute_row = True
            # Could also be VMSS/AKS — skip for AHB purposes (rare to surface).

        # ---- SQL Server on VM (license meters) ----
        elif cat == "Virtual Machines Licenses":
            if sub == "SQL Server Azure Hybrid Benefit":
                workload = "SQL Server on VM"
                ahb_state = "Applied"
                is_license_row = True
                edition = "AHB"
            elif sub.startswith("SQL Server"):
                if "Developer" in sub or "Express" in sub:
                    workload = "SQL Server on VM (free edition)"
                    ahb_state = "N/A"
                    edition = "Developer" if "Developer" in sub else "Express"
                    is_license_row = True
                else:
                    workload = "SQL Server on VM"
                    ahb_state = "Not applied"
                    edition = "Enterprise" if "Enterprise" in sub else (
                              "Standard" if "Standard" in sub else "Other")
                    is_license_row = True
            elif "Red Hat" in sub:
                workload = "RHEL on VM"
                ahb_state = "N/A (BYOS / license-mobility)"
                is_license_row = True
            elif "SUSE" in sub:
                workload = "SUSE on VM"
                ahb_state = "N/A (BYOS / license-mobility)"
                is_license_row = True

        # ---- Azure SQL DB ----
        elif cat == "SQL Database":
            workload = "Azure SQL Database"
            ahb_state = "Review (vCore + license split varies)"
            # Crude AHB hint: if subcategory mentions "SQL License", customer
            # is buying license-included (no AHB).
            if "License" in sub:
                ahb_state = "Not applied (License-included tier)"

        # ---- Azure SQL Managed Instance ----
        elif cat == "SQL Managed Instance":
            workload = "Azure SQL Managed Instance"
            ahb_state = "Review (vCore + license split varies)"
            if "License" in sub:
                ahb_state = "Not applied (License-included tier)"

        if not workload:
            # Update the VM-compute tracker but otherwise ignore.
            continue

        qty  = _to_float(row.get(qty_col))  or 0.0
        cost = _to_float(row.get(cost_col)) or 0.0

        sub_id, rg, rtype, name = _parse_resource_id(rid)
        rid_l = rid.lower()
        agg = aggregations.get(rid_l)
        if agg is None:
            agg = AhbAggregation(
                resource_id=rid,
                name=name or "",
                resource_group=rg or "",
                subscription_id=sub_id or "",
                resource_type=rtype or "",
            )
            aggregations[rid_l] = agg

        if not agg.location:
            agg.location = row.get(reg_col) or ""
        if not agg.currency:
            agg.currency = row.get(cur_col) or ""

        # Track surcharge presence per VM.
        if workload == "Windows on VM" and ahb_state == "Not applied":
            vm_has_surcharge[rid_l] = True

        # Set workload/state on the aggregation. Latest-wins for ahb_state but
        # 'Applied' beats 'Not applied' beats 'Review' beats 'N/A' since that's
        # the order of certainty (Applied is unambiguous; Not applied is
        # observable; Review needs human; N/A means we shouldn't act).
        if not agg.workload:
            agg.workload = workload
            agg.ahb_state = ahb_state
            agg.edition = edition
        else:
            # Prefer the more "actionable" state. Applied > Not applied > Review > N/A.
            order = {"Applied": 4, "Not applied": 3, "Review": 2}
            def rank(s: str) -> int:
                if s.startswith("Applied"):     return 4
                if s.startswith("Not applied"): return 3
                if s.startswith("Review"):      return 2
                return 1
            if rank(ahb_state) > rank(agg.ahb_state):
                agg.ahb_state = ahb_state
                if edition: agg.edition = edition

        agg.total_cost      += cost
        agg.quantity_total  += qty
        if is_license_row:
            agg.license_cost += cost
        if is_compute_row:
            agg.compute_cost += cost

    # Post-pass: synthesize "Windows on VM (AHB applied)" for VMs that had
    # compute rows but never a surcharge row. Need to revisit: we only made
    # an aggregation when we hit a surcharge row, so AHB-applied VMs are NOT
    # currently in `aggregations`. Scan vm_has_compute and add them.
    for rid_l, has_compute in vm_has_compute.items():
        if not has_compute:
            continue
        if vm_has_surcharge.get(rid_l):
            continue  # already in aggregations as "Not applied"
        # Stub aggregation for an AHB-applied (or Linux) VM.
        # To distinguish Linux from Windows-with-AHB we'd need the OS — billing
        # alone can't tell. We mark it 'Applied or N/A' to be honest.
        sub_id, rg, rtype, name = _parse_resource_id(rid_l)
        if rid_l not in aggregations:
            aggregations[rid_l] = AhbAggregation(
                resource_id=rid_l,
                name=name or "",
                resource_group=rg or "",
                subscription_id=sub_id or "",
                resource_type=rtype or "",
                workload="Windows on VM (no surcharge billed)",
                ahb_state="Applied or N/A (Linux or AHB)",
            )

    period_info = _build_period_info(period_start_str, period_end_str,
                                     schema_name=schema["name"])
    elapsed = time.perf_counter() - t0
    print(f"Aggregated {len(aggregations)} AHB-relevant resources in "
          f"{elapsed:.1f}s.")
    return (aggregations, period_info)


def _build_period_info(start: Optional[str], end: Optional[str],
                       *, schema_name: str) -> dict:
    from datetime import datetime
    def _parse(s):
        if not s: return None
        s2 = s.split("T")[0].strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s2, fmt).date()
            except ValueError:
                pass
        return None
    sd, ed = _parse(start), _parse(end)
    days = None
    if sd and ed:
        diff = (ed - sd).days
        if diff > 0:
            days = diff if ed.day == 1 else diff + 1
    if sd and ed and days:
        label = f"{sd.isoformat()} → {ed.isoformat()} ({days} days)"
    elif sd and ed:
        label = f"{sd.isoformat()} → {ed.isoformat()}"
    else:
        label = "unknown period"
    return {"start": sd.isoformat() if sd else None,
            "end":   ed.isoformat() if ed else None,
            "days":  days, "label": label, "schema": schema_name}


# ---------- Build per-resource result rows ----------

def build_results(aggs: dict[str, AhbAggregation]) -> list[dict]:
    out: list[dict] = []
    for agg in aggs.values():
        # Savings if AHB enabled = the license-row cost we observed.
        savings_period = agg.license_cost if agg.ahb_state.startswith("Not applied") else 0.0
        out.append({
            "Workload":            agg.workload,
            "AhbState":            agg.ahb_state,
            "Edition":             agg.edition or None,
            "Name":                agg.name,
            "ResourceGroup":       agg.resource_group,
            "SubscriptionId":      agg.subscription_id,
            "ResourceType":        agg.resource_type,
            "Location":            agg.location,
            "Currency":            agg.currency or None,
            "ComputeCostInPeriod": round(agg.compute_cost, 4) if agg.compute_cost else None,
            "LicenseCostInPeriod": round(agg.license_cost, 4) if agg.license_cost else None,
            "TotalCostInPeriod":   round(agg.total_cost, 4)   if agg.total_cost   else None,
            "PotentialSavingsPeriod": round(savings_period, 4) if savings_period else None,
            "Quantity":            round(agg.quantity_total, 4) if agg.quantity_total else None,
            "ResourceId":          agg.resource_id,
            "Notes":               "; ".join(agg.notes) or None,
        })
    return out


# ---------- CSV resolution ----------

def find_default_usage_csv(script_path: Path) -> Optional[Path]:
    workspace = script_path.parent.parent
    data_dir  = workspace / "data"
    if not data_dir.is_dir():
        return None
    csvs = sorted(data_dir.glob("*.csv"),
                  key=lambda p: p.stat().st_mtime, reverse=True)
    if not csvs:
        return None
    if len(csvs) == 1:
        return csvs[0]
    legacy = [p for p in csvs if p.name.startswith("Detail_BillingProfile_")]
    return legacy[0] if legacy else csvs[0]


# ---------- main ----------

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--usage-csv", help="Detailed billing CSV.")
    parser.add_argument("--export-csv",  default="reports/ahb-report.csv")
    parser.add_argument("--export-xlsx", default="reports/ahb-report.xlsx")
    args = parser.parse_args(argv)

    script_path = Path(__file__).resolve()
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = find_default_usage_csv(script_path)
        if not found:
            print("ERROR: No usage CSV; pass --usage-csv.", file=sys.stderr)
            return 2
        usage_csv = found.resolve()
        print(f"Auto-detected usage CSV: {usage_csv}")
    if not usage_csv.exists():
        print(f"ERROR: Usage CSV not found: {usage_csv}", file=sys.stderr)
        return 2

    aggs, period = aggregate_ahb(usage_csv)
    if not aggs:
        print("\nNo AHB-eligible workloads found in the CSV.")
        return 0

    results = build_results(aggs)
    period_days = period.get("days") or 30
    annualize   = 365.0 / period_days

    # ----- Console summary
    cur = next((r["Currency"] for r in results if r["Currency"]), "")
    by_state: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"count": 0, "license": 0.0, "compute": 0.0, "total": 0.0,
                 "savings": 0.0})
    for r in results:
        k = (r["Workload"], r["AhbState"])
        b = by_state[k]
        b["count"]   += 1
        b["license"] += r["LicenseCostInPeriod"]      or 0
        b["compute"] += r["ComputeCostInPeriod"]      or 0
        b["total"]   += r["TotalCostInPeriod"]        or 0
        b["savings"] += r["PotentialSavingsPeriod"]   or 0

    total_savings_period = sum(r["PotentialSavingsPeriod"] or 0 for r in results)
    total_savings_annual = total_savings_period * annualize

    print()
    print(f"Billing period: {period['label']}")
    print(f"AHB-relevant resources: {len(results)}")
    print()
    print(f"{'Workload':<40} {'State':<40} {'#':>5} {'LicenseCost':>14} {'PotentialSavings':>18}")
    print("-" * 122)
    for (workload, state), b in sorted(by_state.items(),
                                       key=lambda kv: -kv[1]["savings"]):
        print(f"{workload[:40]:<40} {state[:40]:<40} {b['count']:>5} "
              f"{b['license']:>14,.2f} {b['savings']:>18,.2f}")
    print("-" * 122)
    print(f"Total potential savings (period)   : {total_savings_period:>14,.2f} {cur}")
    print(f"Annualized potential savings       : {total_savings_annual:>14,.2f} {cur}/year")

    # Top 10 by potential savings
    not_applied = [r for r in results
                   if (r["PotentialSavingsPeriod"] or 0) > 0]
    not_applied.sort(key=lambda r: -(r["PotentialSavingsPeriod"] or 0))
    if not_applied:
        print("\nTop opportunities (AHB not applied, ranked by period license cost):")
        print(f"  {'Name':<40} {'Workload':<25} {'Edition':<12} "
              f"{'PotentialSavings':>18}")
        for r in not_applied[:15]:
            print(f"  {(r['Name'] or '')[:40]:<40} "
                  f"{(r['Workload'] or '')[:25]:<25} "
                  f"{(r['Edition'] or '')[:12]:<12} "
                  f"{(r['PotentialSavingsPeriod'] or 0):>18,.2f}")
        if len(not_applied) > 15:
            rest = sum(r["PotentialSavingsPeriod"] or 0 for r in not_applied[15:])
            print(f"  ... and {len(not_applied) - 15} more "
                  f"({rest:,.2f} {cur} period savings)")

    # ----- Sort + export
    state_order = {"Not applied": 0, "Not applied (License-included tier)": 1,
                   "Review (vCore + license split varies)": 2,
                   "Applied": 3, "Applied or N/A (Linux or AHB)": 4,
                   "N/A": 5, "N/A (BYOS / license-mobility)": 6}
    results.sort(key=lambda r: (state_order.get(r["AhbState"], 99),
                                -(r["PotentialSavingsPeriod"] or 0),
                                -(r["TotalCostInPeriod"] or 0)))

    out_path = Path(args.export_csv).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)
    print(f"\nReport exported to {out_path}")

    if args.export_xlsx:
        xlsx_path = Path(args.export_xlsx).resolve()
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            write_xlsx(results, xlsx_path,
                       currency=cur, period=period, annualize=annualize,
                       by_state=by_state,
                       total_savings_period=total_savings_period,
                       total_savings_annual=total_savings_annual)
            print(f"Excel report exported to {xlsx_path}")
        except ImportError:
            print("WARNING: openpyxl not installed; skipping Excel export.",
                  file=sys.stderr)
        except PermissionError:
            print(f"WARNING: cannot write {xlsx_path} — is it open in Excel?",
                  file=sys.stderr)
    return 0


def write_xlsx(results, out_path, *, currency, period, annualize,
               by_state, total_savings_period, total_savings_annual) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "AHB"

    bold        = Font(bold=True)
    bold_white  = Font(bold=True, color="FFFFFF")
    big_bold    = Font(bold=True, size=14)
    accent      = Font(bold=True, color="0070C0")
    header_fill = PatternFill("solid", fgColor="305496")
    money_fmt   = "#,##0.00"

    period_label = period.get("label") or "unknown period"
    ws["A1"] = (f"Azure Hybrid Benefit (AHB) scan — "
                f"{len(results)} resources | Period: {period_label}")
    ws["A1"].font = big_bold
    ws.merge_cells("A1:E1")

    # Headline savings
    ws["A3"] = "Potential savings if AHB enabled (period)"
    ws["B3"] = round(total_savings_period, 2); ws["C3"] = currency
    ws["A4"] = "Annualized potential savings"
    ws["B4"] = round(total_savings_annual, 2); ws["C4"] = f"{currency}/year"
    for r in (3, 4):
        ws.cell(row=r, column=1).font = accent
        ws.cell(row=r, column=2).font = accent
        ws.cell(row=r, column=2).number_format = money_fmt

    # Summary by Workload × State
    sum_start = 6
    headers = ["Workload", "State", "Count",
               f"Compute cost ({currency})", f"License cost ({currency})",
               f"Total cost ({currency})", f"Potential savings ({currency})"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=sum_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row = sum_start + 1
    for (workload, state), b in sorted(by_state.items(),
                                       key=lambda kv: -kv[1]["savings"]):
        ws.cell(row=row, column=1, value=workload)
        ws.cell(row=row, column=2, value=state)
        ws.cell(row=row, column=3, value=b["count"])
        ws.cell(row=row, column=4, value=round(b["compute"], 2)).number_format = money_fmt
        ws.cell(row=row, column=5, value=round(b["license"], 2)).number_format = money_fmt
        ws.cell(row=row, column=6, value=round(b["total"],   2)).number_format = money_fmt
        ws.cell(row=row, column=7, value=round(b["savings"], 2)).number_format = money_fmt
        row += 1

    table_start = row + 2

    # Per-resource table
    res_headers = list(results[0].keys())
    for c_idx, h in enumerate(res_headers, start=1):
        cell = ws.cell(row=table_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    money_cols = {"ComputeCostInPeriod", "LicenseCostInPeriod",
                  "TotalCostInPeriod", "PotentialSavingsPeriod"}
    for r_idx, r in enumerate(results, start=table_start + 1):
        for c_idx, h in enumerate(res_headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=r.get(h))
            if h in money_cols and r.get(h) is not None:
                cell.number_format = money_fmt

    # Column widths
    for c_idx, h in enumerate(res_headers, start=1):
        letter = get_column_letter(c_idx)
        max_len = len(h)
        for r in results:
            v = r.get(h)
            if v is None: continue
            max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = ws.cell(row=table_start + 1, column=1)
    ws.auto_filter.ref = (
        f"A{table_start}:{get_column_letter(len(res_headers))}"
        f"{table_start + len(results)}"
    )

    wb.save(out_path)


if __name__ == "__main__":
    sys.exit(main())
