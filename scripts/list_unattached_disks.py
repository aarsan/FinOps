#!/usr/bin/env python3
"""Report unattached Azure managed disks with the customer's negotiated price.

Inventory comes from Azure Resource Graph (a single KQL query across every
subscription the signed-in account can read). Pricing comes from the customer's
detailed billing CSV. Two CSV formats are auto-detected:

  * Legacy 'Detail_BillingProfile_*.csv' (EA / MCA usage detail)
  * FOCUS export ('FocusCost' from Cost Management Exports v2)

Default output: only disks currently in the Unattached state, with monthly
cost, list price, and savings. The headline number is the total monthly
spend the customer would eliminate by deleting all unattached disks.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional


# ----- Tier ladders: (max_provisioned_gb, tier_code) ascending -----
STANDARD_HDD = [
    (32, "S4"), (64, "S6"), (128, "S10"), (256, "S15"), (512, "S20"),
    (1024, "S30"), (2048, "S40"), (4096, "S50"), (8192, "S60"),
    (16384, "S70"), (32767, "S80"),
]
STANDARD_SSD = [
    (4, "E1"), (8, "E2"), (16, "E3"), (32, "E4"), (64, "E6"), (128, "E10"),
    (256, "E15"), (512, "E20"), (1024, "E30"), (2048, "E40"), (4096, "E50"),
    (8192, "E60"), (16384, "E70"), (32767, "E80"),
]
PREMIUM_SSD = [
    (4, "P1"), (8, "P2"), (16, "P3"), (32, "P4"), (64, "P6"), (128, "P10"),
    (256, "P15"), (512, "P20"), (1024, "P30"), (2048, "P40"), (4096, "P50"),
    (8192, "P60"), (16384, "P70"), (32767, "P80"),
]

SKU_PREFIX_TO_PRODUCT = {
    "Standard":    ("Standard HDD Managed Disks", STANDARD_HDD),
    "StandardSSD": ("Standard SSD Managed Disks", STANDARD_SSD),
    "Premium":     ("Premium SSD Managed Disks",  PREMIUM_SSD),
}


@dataclass
class TierLookup:
    tier: Optional[str]
    product: Optional[str]   # CSV meterSubCategory
    meter:   Optional[str]   # CSV meterName, e.g. "P10 LRS Disk"
    note:    Optional[str]


def derive_tier(sku_name: str, size_gb: int) -> TierLookup:
    """Map ('Premium_LRS', 4) → ('P1', 'Premium SSD Managed Disks', 'P1 LRS Disk')."""
    if not sku_name or "_" not in sku_name:
        return TierLookup(None, None, None, f"Unrecognized SKU '{sku_name}'")
    prefix, redundancy = sku_name.split("_", 1)
    info = SKU_PREFIX_TO_PRODUCT.get(prefix)
    if info is None:
        return TierLookup(None, None, None,
                          f"Unsupported SKU '{sku_name}' (Premium SSD v2 / "
                          "Ultra Disks not yet mapped)")
    product, ladder = info
    for cap, tier in ladder:
        if size_gb <= cap:
            return TierLookup(tier, product, f"{tier} {redundancy} Disk", None)
    return TierLookup(None, product, None, f"Size {size_gb} GB exceeds known tiers")


# ---------- CSV → price index ----------

@dataclass
class PriceEntry:
    effective_price:  Optional[float]
    payg_price:       Optional[float]
    contracted_price: Optional[float] = None  # MCA contract rate, before benefits
    currency:         str = ""
    pricing_model:    Optional[str] = None    # OnDemand / Reservation / SavingsPlan / Spot / Standard / Committed
    benefit_name:     Optional[str] = None    # RI / Savings Plan name, when applied
    benefit_type:     Optional[str] = None    # Reservation / SavingsPlan


def _to_float(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


# Two billing CSV schemas are supported. Each schema names the columns we read.
# Legacy: 'Detail_BillingProfile_*.csv' (a.k.a. EA / MCA usage detail)
# FOCUS:  Cost Management 'FocusCost' export (FOCUS 1.0r2 + x_* extensions)
LEGACY_SCHEMA = {
    "name":             "legacy",
    "category":         "meterCategory",
    "subcategory":      "meterSubCategory",
    "meter":            "meterName",
    "region":           "resourceLocation",
    "unit":             "unitOfMeasure",
    "effective_price":  "effectivePrice",
    "list_price":       "payGPrice",
    "contracted_price": "unitPrice",
    "currency":         "pricingCurrency",
    "pricing_model":    "pricingModel",
    "benefit_name":     "benefitName",
    "benefit_type":     None,             # legacy infers from pricingModel
    "resource_id":      "resourceId",
    "quantity":         "quantity",
    "cost":             "costInBillingCurrency",
    "list_cost":        "paygCostInBillingCurrency",
    "date":             "date",
    "category_value":   "Storage",        # value the row's category column should equal
    "unit_pattern":     re.compile(r"^\s*1\s*/\s*Month\s*$"),
}
FOCUS_SCHEMA = {
    "name":             "FOCUS",
    "category":         "ServiceCategory",
    "subcategory":      "x_SkuMeterSubcategory",
    "meter":            "x_SkuMeterName",
    "region":           "RegionId",
    "unit":             "PricingUnit",
    "effective_price":  "x_EffectiveUnitPrice",
    "list_price":       "ListUnitPrice",
    "contracted_price": "ContractedUnitPrice",
    "currency":         "BillingCurrency",
    "pricing_model":    "PricingCategory",  # Standard | Committed | DynamicPricing
    "benefit_name":     "CommitmentDiscountName",
    "benefit_type":     "CommitmentDiscountType",
    "resource_id":      "ResourceId",
    "quantity":         "PricingQuantity",
    "cost":             "EffectiveCost",
    "list_cost":        "ListCost",
    "date":             "ChargePeriodStart",
    "category_value":   "Compute",        # FOCUS classifies managed disks under Compute
    "unit_pattern":     re.compile(r"^\s*Units?\s*/\s*Month\s*$|^\s*1\s*/\s*Month\s*$"),
}


def _detect_schema(fieldnames: list[str]) -> Optional[dict]:
    """Pick the schema whose required columns are all present."""
    for schema in (LEGACY_SCHEMA, FOCUS_SCHEMA):
        cols = {schema[k] for k in (
            "category", "subcategory", "meter", "region",
            "unit", "effective_price", "list_price", "currency",
        )}
        if cols.issubset(fieldnames):
            return schema
    return None


def build_price_index(path: Path) -> dict[tuple[str, str, str], PriceEntry]:
    """Two-stage parse of a (potentially multi-GB) detailed billing CSV.

    1. Stream the file line-by-line and keep only lines that contain both
       'Managed Disks' and ' Disk,' (cheap substring tests reject the 99%+
       of rows that are VMs, networking, etc. without parsing fields).
    2. csv.DictReader over the small surviving subset; auto-detects whether
       the file is a legacy 'Detail_BillingProfile_*' export or a FOCUS
       export (FocusCost), and reads the appropriate columns.

    Returns a dict keyed by (meterSubCategory, meterName, region) lowercased.
    """
    size_mb = path.stat().st_size / (1024 * 1024)
    print(f"Loading customer pricing from '{path}' ({size_mb:,.1f} MB)...")
    t0 = time.perf_counter()

    candidate_lines: list[str] = []
    rows_scanned = 0
    progress_every = 250_000

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        header_line = f.readline()
        if not header_line:
            raise RuntimeError("Usage CSV is empty.")
        candidate_lines.append(header_line)
        for line in f:
            rows_scanned += 1
            if "Managed Disks" in line and " Disk," in line:
                candidate_lines.append(line)
            if rows_scanned % progress_every == 0:
                elapsed = time.perf_counter() - t0
                print(f"  scanned {rows_scanned:,} rows, "
                      f"{len(candidate_lines) - 1:,} disk-candidates so far "
                      f"({elapsed:.1f}s)...")

    elapsed = time.perf_counter() - t0
    print(f"  pre-filter complete: {rows_scanned:,} rows scanned, "
          f"{len(candidate_lines) - 1:,} candidates kept in {elapsed:.1f}s.")

    reader = csv.DictReader(candidate_lines)
    schema = _detect_schema(list(reader.fieldnames or []))
    if schema is None:
        raise RuntimeError(
            "Usage CSV is neither a legacy 'Detail_BillingProfile_*' nor a "
            "FOCUS export. Required columns not found.\n"
            f"  Saw: {sorted(reader.fieldnames or [])[:10]}..."
        )
    print(f"  detected schema: {schema['name']}")

    cat_col   = schema["category"]
    sub_col   = schema["subcategory"]
    meter_col = schema["meter"]
    reg_col   = schema["region"]
    unit_col  = schema["unit"]
    eff_col   = schema["effective_price"]
    list_col  = schema["list_price"]
    cur_col   = schema["currency"]
    cat_value = schema["category_value"]
    unit_re   = schema["unit_pattern"]
    # Optional columns — may not exist on older exports.
    contract_col = schema.get("contracted_price") if schema.get("contracted_price") in (reader.fieldnames or []) else None
    pm_col       = schema.get("pricing_model")    if schema.get("pricing_model")    in (reader.fieldnames or []) else None
    bn_col       = schema.get("benefit_name")     if schema.get("benefit_name")     in (reader.fieldnames or []) else None
    bt_col       = schema.get("benefit_type")     if schema.get("benefit_type")     in (reader.fieldnames or []) else None

    index: dict[tuple[str, str, str], PriceEntry] = {}
    matched = 0

    for row in reader:
        if row.get(cat_col) != cat_value:
            continue
        sub = row.get(sub_col) or ""
        if not sub.endswith("Managed Disks"):
            continue
        meter = row.get(meter_col) or ""
        if not meter.endswith(" Disk"):
            continue
        if not unit_re.match(row.get(unit_col) or ""):
            continue

        matched += 1
        region = (row.get(reg_col) or "").lower()
        key = (sub.lower(), meter.lower(), region)
        existing = index.get(key)
        if existing is not None and existing.effective_price is not None:
            continue

        index[key] = PriceEntry(
            effective_price=_to_float(row.get(eff_col)),
            payg_price=_to_float(row.get(list_col)),
            contracted_price=_to_float(row.get(contract_col)) if contract_col else None,
            currency=row.get(cur_col) or "",
            pricing_model=(row.get(pm_col) or None) if pm_col else None,
            benefit_name=(row.get(bn_col) or None) if bn_col else None,
            benefit_type=(row.get(bt_col) or None) if bt_col else None,
        )

    elapsed = time.perf_counter() - t0
    print(f"Indexed {len(index)} unique disk price points from {matched} matching "
          f"rows ({rows_scanned:,} total rows scanned in {elapsed:.1f}s).")
    return index


# ---------- Resource Graph → disk inventory ----------

def query_disks_via_resource_graph(
    *, all_states: bool = False, tenant_id: Optional[str] = None,
) -> list[dict]:
    """Run a KQL query through Azure Resource Graph for managed disks.

    Returns a list of dicts with keys: id, name, resourceGroup, subscriptionId,
    location, skuName, sizeGB, diskState, timeCreated. By default filters to
    disks currently in the Unattached state.
    """
    try:
        from azure.identity import (
            ChainedTokenCredential,
            DefaultAzureCredential,
            InteractiveBrowserCredential,
        )
        from azure.mgmt.resourcegraph import ResourceGraphClient
        from azure.mgmt.resourcegraph.models import (
            QueryRequest, QueryRequestOptions, ResultFormat,
        )
        from azure.mgmt.subscription import SubscriptionClient
    except ImportError as exc:
        raise RuntimeError(
            "Azure SDK packages not installed. Install with:\n"
            "  pip install -r scripts/requirements.txt"
        ) from exc

    # Try every silent credential source the SDK supports (env vars, managed
    # identity, VS Code / Visual Studio sign-in, az / azd CLI), then fall
    # back to launching a browser window for interactive sign-in. The customer
    # does NOT need the Azure CLI installed.
    silent_kwargs = {"additionally_allowed_tenants": ["*"]}
    interactive_kwargs: dict[str, object] = {}
    if tenant_id:
        silent_kwargs["tenant_id"] = tenant_id
        interactive_kwargs["tenant_id"] = tenant_id

    credential = ChainedTokenCredential(
        DefaultAzureCredential(**silent_kwargs),
        InteractiveBrowserCredential(**interactive_kwargs),
    )

    # Enumerate every subscription this credential can see.
    print("Enumerating accessible Azure subscriptions...")
    sub_client = SubscriptionClient(credential)
    sub_ids = [s.subscription_id for s in sub_client.subscriptions.list()
               if s.subscription_id]
    if not sub_ids:
        raise RuntimeError(
            "No subscriptions are visible to this account. "
            "Pass --tenant-id <tenant-id> to sign in to the right directory."
        )
    print(f"  found {len(sub_ids)} subscription(s).")

    state_filter = "" if all_states else \
        "| where properties.diskState == 'Unattached'\n    "
    kql = f"""
    resources
    | where type =~ 'microsoft.compute/disks'
    {state_filter}| project id, name, resourceGroup, subscriptionId, location,
              skuName=tostring(sku.name),
              sizeGB=toint(properties.diskSizeGB),
              diskState=tostring(properties.diskState),
              timeCreated=tostring(properties.timeCreated)
    """

    arg_client = ResourceGraphClient(credential)
    all_rows: list[dict] = []
    skip_token: Optional[str] = None
    page = 0

    label = "all" if all_states else "unattached"
    print(f"Querying Resource Graph for {label} managed disks...")
    while True:
        page += 1
        opts = QueryRequestOptions(
            top=1000,
            skip_token=skip_token,
            result_format=ResultFormat.OBJECT_ARRAY,
            allow_partial_scopes=True,
        )
        req = QueryRequest(subscriptions=sub_ids, query=kql, options=opts)
        resp = arg_client.resources(req)
        rows = list(resp.data) if resp.data else []
        all_rows.extend(rows)
        print(f"  page {page}: +{len(rows)} (running total {len(all_rows)})")
        skip_token = resp.skip_token
        if not skip_token:
            break

    print(f"Resource Graph returned {len(all_rows)} disk(s).")
    return all_rows


# ---------- CSV utilities ----------

def find_default_usage_csv(script_path: Path) -> Optional[Path]:
    """Find the usage CSV under <workspace>/data/.

    Resolution order:
      1. The single *.csv file in data/, if there is exactly one.
      2. The most recent Detail_BillingProfile_*.csv (legacy naming).
      3. The most recent *.csv overall.
    Returns None if data/ is missing or contains no CSVs.
    """
    workspace = script_path.parent.parent
    data_dir  = workspace / "data"
    if not data_dir.is_dir():
        return None

    csvs = sorted(
        data_dir.glob("*.csv"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not csvs:
        return None
    if len(csvs) == 1:
        return csvs[0]

    # Multiple CSVs: prefer Detail_BillingProfile_*.csv if present, otherwise
    # the most recent file.
    legacy = [p for p in csvs if p.name.startswith("Detail_BillingProfile_")]
    return legacy[0] if legacy else csvs[0]


# ---------- Verify mode ----------

def _resolve_usage_csv(args, script_path: Path) -> Path:
    """Shared usage-CSV resolution for both report and verify modes."""
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = find_default_usage_csv(script_path)
        if not found:
            raise SystemExit(
                "ERROR: No usage CSV provided and no *.csv found in "
                "<workspace>/data/. Pass --usage-csv."
            )
        usage_csv = found.resolve()
        print(f"Auto-detected usage CSV: {usage_csv}")
    if not usage_csv.exists():
        raise SystemExit(f"ERROR: Usage CSV not found: {usage_csv}")
    return usage_csv


def _stream_disk_rows_for_resource(path: Path, resource_id: str) -> tuple[dict, list[dict]]:
    """Stream the CSV, return (schema, [matching rows for resource_id])."""
    needle = resource_id.lower()
    print(f"Searching CSV for billing rows matching {resource_id} ...")
    t0 = time.perf_counter()

    candidate_lines: list[str] = []
    rows_scanned = 0
    progress_every = 250_000

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        header_line = f.readline()
        if not header_line:
            raise RuntimeError("Usage CSV is empty.")
        candidate_lines.append(header_line)
        for line in f:
            rows_scanned += 1
            if needle in line.lower():
                candidate_lines.append(line)
            if rows_scanned % progress_every == 0:
                elapsed = time.perf_counter() - t0
                print(f"  scanned {rows_scanned:,} rows, "
                      f"{len(candidate_lines) - 1:,} candidates so far "
                      f"({elapsed:.1f}s)...")

    elapsed = time.perf_counter() - t0
    print(f"  pre-filter complete: {rows_scanned:,} rows scanned, "
          f"{len(candidate_lines) - 1:,} candidates in {elapsed:.1f}s.")

    if len(candidate_lines) == 1:
        return ({}, [])

    reader = csv.DictReader(candidate_lines)
    schema = _detect_schema(list(reader.fieldnames or []))
    if schema is None:
        raise RuntimeError("CSV schema not recognized (legacy or FOCUS).")

    rid_col = schema.get("resource_id")
    if not rid_col or rid_col not in (reader.fieldnames or []):
        raise RuntimeError(
            f"CSV is missing the resource-id column ({rid_col!r})."
        )

    rows = [r for r in reader if (r.get(rid_col) or "").lower() == needle]
    print(f"  exact resourceId matches: {len(rows)}")
    return (schema, rows)


def _run_verify(args, script_path: Path) -> int:
    needle = args.verify

    # ----- 1. Look up the disk in Resource Graph (by name OR full id)
    try:
        from azure.identity import (
            ChainedTokenCredential, DefaultAzureCredential,
            InteractiveBrowserCredential,
        )
        from azure.mgmt.resourcegraph import ResourceGraphClient
        from azure.mgmt.resourcegraph.models import (
            QueryRequest, QueryRequestOptions, ResultFormat,
        )
        from azure.mgmt.subscription import SubscriptionClient
    except ImportError as exc:
        print(f"\nERROR: Azure SDK packages not installed: {exc}", file=sys.stderr)
        return 1

    silent_kwargs = {"additionally_allowed_tenants": ["*"]}
    interactive_kwargs: dict[str, object] = {}
    if args.tenant_id:
        silent_kwargs["tenant_id"] = args.tenant_id
        interactive_kwargs["tenant_id"] = args.tenant_id
    credential = ChainedTokenCredential(
        DefaultAzureCredential(**silent_kwargs),
        InteractiveBrowserCredential(**interactive_kwargs),
    )

    print("Enumerating accessible Azure subscriptions...")
    sub_ids = [s.subscription_id for s in SubscriptionClient(credential)
               .subscriptions.list() if s.subscription_id]
    if not sub_ids:
        print("ERROR: no subscriptions visible to this account.", file=sys.stderr)
        return 1

    if needle.startswith("/subscriptions/"):
        kql_filter = f"| where id =~ '{needle}'"
    else:
        kql_filter = f"| where name =~ '{needle}'"
    kql = f"""
    resources
    | where type =~ 'microsoft.compute/disks'
    {kql_filter}
    | project id, name, resourceGroup, subscriptionId, location,
              skuName=tostring(sku.name),
              sizeGB=toint(properties.diskSizeGB),
              diskState=tostring(properties.diskState),
              timeCreated=tostring(properties.timeCreated)
    """
    arg_client = ResourceGraphClient(credential)
    resp = arg_client.resources(QueryRequest(
        subscriptions=sub_ids, query=kql,
        options=QueryRequestOptions(top=10,
                                    result_format=ResultFormat.OBJECT_ARRAY,
                                    allow_partial_scopes=True),
    ))
    matches = list(resp.data) if resp.data else []
    if not matches:
        print(f"\nNo disk in Resource Graph matches '{needle}'.", file=sys.stderr)
        return 2
    if len(matches) > 1:
        print(f"\n{len(matches)} disks match '{needle}'. Pass the full "
              "resource id (--verify /subscriptions/...) to disambiguate.",
              file=sys.stderr)
        for d in matches:
            print(f"  - {d['id']}", file=sys.stderr)
        return 2

    disk = matches[0]
    print()
    print("=" * 70)
    print("Resource Graph record")
    print("=" * 70)
    for k in ("name", "resourceGroup", "subscriptionId", "location",
              "skuName", "sizeGB", "diskState", "timeCreated", "id"):
        print(f"  {k:<15}: {disk.get(k)}")

    tier_info = derive_tier(disk.get("skuName") or "",
                            int(disk.get("sizeGB") or 0))
    print(f"  derivedTier    : {tier_info.tier}")
    print(f"  derivedMeter   : {tier_info.meter}")
    print(f"  derivedProduct : {tier_info.product}")

    # ----- 2. Pull every CSV row for this resourceId
    usage_csv = _resolve_usage_csv(args, script_path)
    schema, rows = _stream_disk_rows_for_resource(usage_csv, disk["id"])

    print()
    print("=" * 70)
    print(f"CSV billing rows ({schema.get('name', '?')} schema)")
    print("=" * 70)
    if not rows:
        print("No billing rows reference this resource id. Either the disk "
              "wasn't billed during the export period, the export covers a "
              "different scope, or attachment-state filtering excluded it.")
        return 0

    meter_col = schema["meter"]
    sub_col   = schema["subcategory"]
    reg_col   = schema["region"]
    eff_col   = schema["effective_price"]
    list_col  = schema["list_price"]
    qty_col   = schema.get("quantity")
    cost_col  = schema.get("cost")
    lcost_col = schema.get("list_cost")
    date_col  = schema.get("date")
    cur_col   = schema["currency"]
    bn_col    = schema.get("benefit_name")

    # Per-day roll-up by (meter, region, date).
    by_day: dict[tuple, dict] = {}
    for r in rows:
        key = (r.get(meter_col) or "", r.get(reg_col) or "",
               (r.get(date_col) or "").split("T")[0])
        agg = by_day.setdefault(key, {
            "meter": key[0], "region": key[1], "date": key[2],
            "qty": 0.0, "cost": 0.0, "list_cost": 0.0,
            "rows": 0,
            "eff_price": _to_float(r.get(eff_col)),
            "list_price": _to_float(r.get(list_col)),
            "currency": r.get(cur_col),
            "benefit": (r.get(bn_col) or None) if bn_col else None,
        })
        agg["qty"]       += _to_float(r.get(qty_col))  or 0.0
        agg["cost"]      += _to_float(r.get(cost_col)) or 0.0
        if lcost_col:
            agg["list_cost"] += _to_float(r.get(lcost_col)) or 0.0
        agg["rows"]      += 1

    # Print sorted by date.
    print(f"  {'date':<12} {'meter':<28} {'region':<14} "
          f"{'qty':>10} {'cost':>12} {'list':>12}  benefit")
    print("  " + "-" * 100)
    for k in sorted(by_day):
        a = by_day[k]
        print(f"  {a['date']:<12} {a['meter'][:28]:<28} {a['region']:<14} "
              f"{a['qty']:>10.4f} {a['cost']:>12.4f} {a['list_cost']:>12.4f}  "
              f"{a['benefit'] or ''}")

    # Totals.
    total_qty  = sum(a["qty"]       for a in by_day.values())
    total_cost = sum(a["cost"]      for a in by_day.values())
    total_list = sum(a["list_cost"] for a in by_day.values())
    cur = next((a["currency"] for a in by_day.values() if a["currency"]), "")
    print("  " + "-" * 100)
    print(f"  {'TOTAL':<56} {total_qty:>10.4f} {total_cost:>12.4f} "
          f"{total_list:>12.4f}  {cur}")

    # ----- 3. Compare to what the report's join logic would have used
    print()
    print("=" * 70)
    print("Reconciliation")
    print("=" * 70)
    if tier_info.product and tier_info.meter:
        any_eff   = next((a["eff_price"]  for a in by_day.values()
                          if a["eff_price"]  is not None), None)
        any_list  = next((a["list_price"] for a in by_day.values()
                          if a["list_price"] is not None), None)
        # Did the report's tier derivation match the meter actually billed?
        billed_meters = {a["meter"] for a in by_day.values()}
        if tier_info.meter in billed_meters:
            print(f"  Derived meter '{tier_info.meter}' matches billed meter(s) ✔")
        else:
            print(f"  WARNING: derived meter '{tier_info.meter}' NOT in "
                  f"billed meters {sorted(billed_meters)}. "
                  "Disk may have been resized mid-period.")
        if any_eff is not None:
            print(f"  Per-unit effective price (CSV): {any_eff:.4f} {cur}")
            print(f"  Per-unit list price      (CSV): "
                  f"{any_list:.4f} {cur}" if any_list is not None else
                  "  Per-unit list price      (CSV): n/a")
            print(f"  Implied monthly cost (price × 1): {any_eff:.2f} {cur}")
        if total_cost > 0:
            print(f"  Actual cost across all CSV rows  : {total_cost:.2f} {cur}")
            print(f"  Actual list cost (PAYG)          : {total_list:.2f} {cur}")
            if total_qty > 0:
                # Approximate per-unit by dividing aggregate cost / quantity
                derived_unit = total_cost / total_qty
                print(f"  Derived unit cost (cost / qty)   : {derived_unit:.4f} {cur}")
    return 0


# ---------- main ----------

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--usage-csv",
        help="Detailed billing CSV. If omitted, the script picks the single "
             "CSV under <workspace>/data/ (or the most recent if there are "
             "multiple).",
    )
    parser.add_argument(
        "--export-csv",
        default="reports/unattached-disks-report.csv",
        help="Output CSV path (default: %(default)s).",
    )
    parser.add_argument(
        "--export-xlsx",
        default="reports/unattached-disks-report.xlsx",
        help="Output Excel path (default: %(default)s). "
             "Pass an empty string to skip.",
    )
    parser.add_argument(
        "--all-disks",
        action="store_true",
        help="Include disks in any state, not just Unattached.",
    )
    parser.add_argument(
        "--tenant-id",
        help="Azure AD tenant ID to authenticate against. If omitted, the "
             "default credential's home tenant is used.",
    )
    parser.add_argument(
        "--verify",
        metavar="DISK_NAME_OR_ID",
        help="Reconcile a single disk against the CSV: prints the Resource "
             "Graph record, every CSV billing row for that resourceId, and a "
             "month-over-month roll-up. Skips the normal report.",
    )
    args = parser.parse_args(argv)

    script_path = Path(__file__).resolve()

    # ----- Verify mode: reconcile one disk against the CSV
    if args.verify:
        return _run_verify(args, script_path)

    # ----- 1. Disk inventory from Resource Graph
    try:
        disks = query_disks_via_resource_graph(
            all_states=args.all_disks, tenant_id=args.tenant_id,
        )
    except Exception as exc:  # noqa: BLE001 — surface SDK / auth errors plainly
        print(f"\nERROR: Resource Graph query failed: {exc}", file=sys.stderr)
        print("If sign-in failed, retry with --tenant-id <tenant-id>.",
              file=sys.stderr)
        return 1

    if not disks:
        label = "disks" if args.all_disks else "unattached disks"
        print(f"\nNo {label} found.")
        return 0

    # ----- 2. Price index from billing CSV
    if args.usage_csv:
        usage_csv = Path(args.usage_csv).resolve()
    else:
        found = find_default_usage_csv(script_path)
        if not found:
            print("\nERROR: No usage CSV provided and no *.csv found in "
                  "<workspace>/data/. Pass --usage-csv.", file=sys.stderr)
            return 2
        usage_csv = found.resolve()
        print(f"\nAuto-detected usage CSV: {usage_csv}")

    if not usage_csv.exists():
        print(f"\nERROR: Usage CSV not found: {usage_csv}", file=sys.stderr)
        return 2

    price_index = build_price_index(usage_csv)

    # ----- 3. Marry inventory to pricing
    results: list[dict] = []
    no_price = 0
    for d in disks:
        sku  = d.get("skuName") or ""
        size = int(d.get("sizeGB") or 0)
        loc  = (d.get("location") or "").lower()
        tier_info = derive_tier(sku, size)

        effective_price = payg_price = contracted_price = None
        customer_monthly = list_monthly = monthly_savings = None
        currency = ""
        benefit = "List price"
        benefit_discount_pct: Optional[float] = None
        contract_discount_pct: Optional[float] = None
        if tier_info.product and tier_info.meter:
            entry = price_index.get(
                (tier_info.product.lower(), tier_info.meter.lower(), loc)
            )
            if entry and entry.effective_price is not None:
                effective_price  = round(entry.effective_price, 4)
                customer_monthly = round(entry.effective_price, 2)
                if entry.contracted_price is not None:
                    contracted_price = round(entry.contracted_price, 4)
                if entry.payg_price is not None:
                    payg_price      = round(entry.payg_price, 4)
                    list_monthly    = round(entry.payg_price, 2)
                    monthly_savings = round(entry.payg_price - entry.effective_price, 2)
                    if entry.payg_price > 0:
                        contract_discount_pct = round(
                            (entry.payg_price - (entry.contracted_price
                                                 if entry.contracted_price is not None
                                                 else entry.effective_price))
                            / entry.payg_price * 100, 2
                        )
                currency = entry.currency

                # Reason effective < list. Priority: explicit benefit > contracted > list.
                if entry.benefit_name:
                    btype = entry.benefit_type or entry.pricing_model or "Benefit"
                    benefit = f"{btype}: {entry.benefit_name}"
                    if (entry.payg_price and entry.payg_price > 0
                            and entry.effective_price is not None):
                        benefit_discount_pct = round(
                            (entry.payg_price - entry.effective_price)
                            / entry.payg_price * 100, 2
                        )
                elif (entry.contracted_price is not None and entry.payg_price is not None
                      and entry.contracted_price < entry.payg_price - 1e-9):
                    benefit = "MCA negotiated rate"
                elif (entry.payg_price is not None
                      and entry.effective_price < entry.payg_price - 1e-9):
                    # Some legacy exports leave unitPrice == payGPrice but discount lives in effectivePrice.
                    benefit = "Negotiated rate"
                else:
                    benefit = "List price"
            else:
                no_price += 1
        else:
            no_price += 1

        results.append({
            "Name":                 d.get("name"),
            "ResourceGroup":        d.get("resourceGroup"),
            "SubscriptionId":       d.get("subscriptionId"),
            "Location":             d.get("location"),
            "DiskState":            d.get("diskState"),
            "SizeGB":               size or None,
            "Sku":                  sku or None,
            "Tier":                 tier_info.tier,
            "Currency":             currency or None,
            "EffectivePrice":       effective_price,
            "ContractedPrice":      contracted_price,
            "PayGPrice":            payg_price,
            "Benefit":              benefit,
            "BenefitDiscountPct":   benefit_discount_pct,    # vs list, if a named benefit was applied
            "ContractDiscountPct":  contract_discount_pct,   # vs list (sum of contract + benefit)
            "CustomerMonthlyCost":  customer_monthly,
            "ListMonthlyCost":      list_monthly,
            "MonthlySavings":       monthly_savings,
        })

    if no_price:
        print(f"\nNote: {no_price} disk(s) had no matching price in the CSV "
              "(likely Premium SSD v2 / Ultra, or a region not present in the "
              "CSV's billing period).")

    results.sort(key=lambda r: -(r["CustomerMonthlyCost"] or 0))

    total_customer = sum(r["CustomerMonthlyCost"] or 0 for r in results)
    total_list     = sum(r["ListMonthlyCost"]     or 0 for r in results)
    total_savings  = sum(r["MonthlySavings"]      or 0 for r in results)
    cur = next((r["Currency"] for r in results if r["Currency"]), "")

    label = "managed disks" if args.all_disks else "unattached disks"
    print()
    print(f"Total {label}: {len(results)}")
    print(f"Customer monthly cost (savings if all deleted)  : {total_customer:,.2f} {cur}")
    print(f"List monthly cost                               : {total_list:,.2f} {cur}")
    print(f"Additional savings vs list (negotiated discount): {total_savings:,.2f} {cur}")

    out_path = Path(args.export_csv).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)
    print(f"\nReport exported to {out_path}")

    if args.export_xlsx:
        xlsx_path = Path(args.export_xlsx).resolve()
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            write_xlsx(results, xlsx_path, currency=cur,
                       total_customer=total_customer,
                       total_list=total_list,
                       total_savings=total_savings,
                       label=label)
            print(f"Excel report exported to {xlsx_path}")
        except ImportError:
            print("WARNING: openpyxl not installed; skipping Excel export.",
                  file=sys.stderr)
        except PermissionError:
            print(f"WARNING: cannot write {xlsx_path} — is it open in Excel? "
                  f"Close it or pass --export-xlsx <other-path>.",
                  file=sys.stderr)

    return 0


def write_xlsx(
    results: list[dict],
    out_path: Path,
    *,
    currency: str,
    total_customer: float,
    total_list: float,
    total_savings: float,
    label: str,
) -> None:
    """Single-sheet workbook: title, totals, then a disk table."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Disks"

    bold        = Font(bold=True)
    bold_white  = Font(bold=True, color="FFFFFF")
    big_bold    = Font(bold=True, size=14)
    accent      = Font(bold=True, color="0070C0")
    header_fill = PatternFill("solid", fgColor="305496")
    money_fmt   = '#,##0.00'

    ws["A1"] = f"Unattached Disks — {len(results)} disks ({label})"
    ws["A1"].font = big_bold
    ws.merge_cells("A1:D1")

    ws["A3"] = "Total customer monthly cost"
    ws["B3"] = round(total_customer, 2)
    ws["C3"] = currency
    ws["A4"] = "Total list monthly cost"
    ws["B4"] = round(total_list, 2)
    ws["C4"] = currency
    ws["A5"] = "Total monthly savings if all deleted"
    ws["B5"] = round(total_customer, 2)
    ws["C5"] = currency
    ws["A6"] = "Additional savings vs list (negotiated discount)"
    ws["B6"] = round(total_savings, 2)
    ws["C6"] = currency

    for r in (3, 4, 5, 6):
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=2).number_format = money_fmt
    ws.cell(row=5, column=1).font = accent
    ws.cell(row=5, column=2).font = accent

    table_start = 8
    headers = list(results[0].keys())
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=c_idx, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_cols = {
        "EffectivePrice", "ContractedPrice", "PayGPrice",
        "CustomerMonthlyCost", "ListMonthlyCost", "MonthlySavings",
    }
    pct_cols = {"BenefitDiscountPct", "ContractDiscountPct"}
    for r_idx, row in enumerate(results, start=table_start + 1):
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            if h in money_cols and row[h] is not None:
                cell.number_format = money_fmt
            elif h in pct_cols and row[h] is not None:
                cell.number_format = '0.00"%"'

    for c_idx, h in enumerate(headers, start=1):
        letter = get_column_letter(c_idx)
        max_len = len(h)
        for row in results:
            v = row[h]
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = ws.cell(row=table_start + 1, column=1)
    ws.auto_filter.ref = (
        f"A{table_start}:{get_column_letter(len(headers))}"
        f"{table_start + len(results)}"
    )

    wb.save(out_path)


if __name__ == "__main__":
    sys.exit(main())
